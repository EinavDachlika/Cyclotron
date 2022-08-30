from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import datetime
from PIL import Image, ImageTk, ImageFont
import mysql.connector
from mysql.connector import Error
from datetime import datetime , date, timedelta
from tkcalendar import Calendar, DateEntry
from openpyxl import *
from openpyxl.styles import *
from pathlib import Path
import math
from numpy import log as ln
import webbrowser
from operator import itemgetter

# from importlib import reload


root = Tk()
#root.geometry("300x300")

root.title("Settings")

#defult font
root.option_add("*Font", "Helvetica")


#general
label_font = ('Helvetica',26, 'bold')
label_font_flag_on_page = ('Helvetica 12 bold underline')
label_font_flag = ('Helvetica 12')
sub_label_font = ('Helvetica',18, 'bold')
label_color = '#034672'
red_color =  '#f5bfbf'

# connect to MySqL
try:

    # # Maor local DB Mysql
    # db = mysql.connector.connect(
    #     host="localhost",
    #     port=3308,
    #     user="root",
    #     password="root",
    #     database="cyclotron")

    # Einav local DB-Mysql
    db = mysql.connector.connect(
      host="localhost",
      user="root",
      password="Cyclotron2022@?%",
      database= "cyclotron")

    if db.is_connected():
        # db_Info = db.get_server_info()
        # print("Connected to MySQL Server version ", db_Info)
        dbCursor = db.cursor(buffered=True)
        # dbCursor.execute("select database();")
        # record = dbCursor.fetchone()
        # print("You're connected to database: ", record)

except Error as e:
    print("Error while connecting to MySQL", e)

cursor = db.cursor()
#toolbar function
def work_plan_page():
    WorkPlanFrame.pack(fill=X)
    moduleSettingsFrame.forget()
    materialSettingsFrame.forget()
    hospitalFrame.forget()
    cycloSettingsFrame.forget()
    batchFrame.forget()

def batch_page():
    batchFrame.pack(fill=X)
    moduleSettingsFrame.forget()
    materialSettingsFrame.forget()
    hospitalFrame.forget()
    cycloSettingsFrame.forget()
    WorkPlanFrame.forget()
##################### toolbar #####################

toolbarbgcolor = "white"
toolbar = Frame(root, bg=toolbarbgcolor)
toolbar.grid(sticky='nesw')

# logo - toolbar
LogoImagePath = Image.open("LogoImage.png")
LogoImageResize = LogoImagePath.resize((120, 57),Image.ANTIALIAS)
LogoImage = ImageTk.PhotoImage(LogoImageResize)
Label(toolbar,image=LogoImage).pack(side=LEFT,padx=10,pady=6)

# work plan button - toolbar
workPlanButton = Button(toolbar, text="Work Plans",font='Helvetica 11',  command=lambda: work_plan_page())
workPlanButton.pack(side=LEFT,padx=10,pady=3)


# Orders button - toolbar
ordersButton = Button (toolbar, text="Orders", font='Helvetica 11')
ordersButton.pack(side=LEFT,padx=10,pady=3)

# Batches button - toolbar
ordersButton = Button (toolbar, text="Batches", font='Helvetica 11',command=lambda: batch_page())
ordersButton.pack(side=LEFT,padx=10,pady=3)


# Reports button - toolbar
reportsButton = Button (toolbar, text="Reports", font='Helvetica 11')
reportsButton.pack(side=LEFT,padx=10,pady=3)

# settings Icon - toolbar

settingsIcon = Image.open("gearIcon.png")
resizedSettingsIcon = settingsIcon.resize((35,35), Image.ANTIALIAS)
imgSettings = ImageTk.PhotoImage(resizedSettingsIcon)
# Button(toolbar, image=imgSettings, borderwidth=0).pack(side=RIGHT,padx=10,pady=3)
mbtn = Menubutton(toolbar, image=imgSettings, borderwidth=0)
mbtn.pack(side=RIGHT,padx=10,pady=3)
mbtn.menu = Menu(mbtn, tearoff = 0)
mbtn["menu"] = mbtn.menu
selected_settings_option = StringVar()

def menu_item_selected(label):
    if label == 'Cyclotron':
        cycloSettingsFrame.pack(fill=X)
        moduleSettingsFrame.forget()
        materialSettingsFrame.forget()
        hospitalFrame.forget()
        WorkPlanFrame.forget()
        batchFrame.forget()


    elif label == 'Module':
        moduleSettingsFrame.pack(fill=X)
        cycloSettingsFrame.forget()
        materialSettingsFrame.forget()
        hospitalFrame.forget()
        WorkPlanFrame.forget()
        batchFrame.forget()

    elif label == 'Hospital':
        hospitalFrame.pack(fill=X)
        cycloSettingsFrame.forget()
        materialSettingsFrame.forget()
        moduleSettingsFrame.forget()
        WorkPlanFrame.forget()
        batchFrame.forget()

    else:
        materialSettingsFrame.pack(fill=X)
        cycloSettingsFrame.forget()
        moduleSettingsFrame.forget()
        hospitalFrame.forget()
        WorkPlanFrame.forget()
        batchFrame.forget()


selected_settings_option.trace("w", menu_item_selected)

mbtn.menu.add_radiobutton(label="Cyclotron", command= lambda: menu_item_selected("Cyclotron"))
mbtn.menu.add_radiobutton(label="Module", command= lambda: menu_item_selected("Module"))
mbtn.menu.add_radiobutton(label="Material", command= lambda: menu_item_selected("Material"))
mbtn.menu.add_radiobutton(label="Hospital", command= lambda: menu_item_selected("Hospital"))



# print(mbtn.selection_get())
toolbar.pack(side=TOP, fill=X)

toolbar.grid_columnconfigure(1, weight=1)


dict_input_column = { 'hospital':('Name', 'Fixed_activity_level', 'Transport_time_min', 'Transport_time_max') ,
                       'resourcecyclotron':('version', 'capacity', 'constant_efficiency', 'description') ,
                      'resourcemodule': ('version', 'capacity', 'description' ) ,
                      'material':('materialName'),
                      'batch': ('TargetCurrentLB','DecayCorrected_TTA' , 'EOS_activity')}

#Einav
query_index_col = """select 
        col.table_name as 'table',
        col.ordinal_position as col_id,
        col.column_name as column_name
        from information_schema.columns col
        where  TABLE_SCHEMA='cyclotron'
         order by col.table_name, col.ordinal_position """
cursor.execute(query_index_col)
dic_metadata = cursor.fetchall()
#end Einav

dataType_col = """SELECT table_name,column_name, DATA_TYPE
                FROM INFORMATION_SCHEMA.COLUMNS where TABLE_SCHEMA='cyclotron' """

cursor.execute(dataType_col)
dataType_col_list = cursor.fetchall()


table_pk_list = """select 
        # sta.index_name as pk_name,
        tab.table_name,
        sta.column_name,
        sta.seq_in_index as column_id
    from information_schema.tables as tab
    inner join information_schema.statistics as sta
            on sta.table_schema = tab.table_schema
            and sta.table_name = tab.table_name
            and sta.index_name = 'primary'
    where tab.table_schema = 'cyclotron'
        and tab.table_type = 'BASE TABLE'
    order by tab.table_name,
        column_id;"""
cursor.execute(table_pk_list)
table_pk_list = cursor.fetchall()


fk_query = """select 
       col.table_name as 'table',
       kcu.constraint_name as fk_constraint_name,
       # col.ordinal_position as col_id,
       # col.column_name as column_name,
       # case when kcu.referenced_table_schema is null
       #      then null
       #      else '>-' end as rel,
       kcu.referenced_table_name as primary_table,
       kcu.referenced_column_name as pk_column_name
from information_schema.columns col
join information_schema.tables tab
     on col.table_schema = tab.table_schema
     and col.table_name = tab.table_name
left join information_schema.key_column_usage kcu
     on col.table_schema = kcu.table_schema
     and col.table_name = kcu.table_name
     and col.column_name = kcu.column_name
     and kcu.referenced_table_schema is not null
where col.table_schema not in('information_schema','sys',
                              'mysql', 'performance_schema')
      and tab.table_type = 'BASE TABLE'
--    and fks.constraint_schema = 'cyclotron'
      and col.table_schema = 'cyclotron'
      and kcu.constraint_name is not null
order by col.table_schema,
         col.table_name,
         col.ordinal_position;"""
cursor.execute(fk_query)
fk = cursor.fetchall()
# print(fk)
# print([data for data in fk if data[3]=='idworkplan'])

query = "SELECT materialName,idmaterial FROM material"
cursor.execute(query)
material_options_list = cursor.fetchall()


def NOT_NULL_DataType_col(table_name):
    # column that define as NOT NULL in db
    # query = "select TABLE_NAME, COLUMN_NAME, IS_NULLABLE from information_schema.COLUMNS where TABLE_SCHEMA='cyclotron'and IS_NULLABLE='NO'order by ordinal_position "

    query = """SELECT table_name,column_name, DATA_TYPE , IS_NULLABLE
                    FROM INFORMATION_SCHEMA.COLUMNS where TABLE_SCHEMA='cyclotron' and table_name= '"""
    query = query+table_name +"'"
    cursor.execute(query)
    data = cursor.fetchall()
    return data


def error_message(text):
    messagebox.showerror("Error",text)

def warning_message(text):
    messagebox.showwarning("Warning",text)

def YES_NO_message(title_tab, text):
    return messagebox.askyesno(title_tab,text)


#algorithm functions
lamda = ln(2) / 109.6
max_activity_batch=7300


def sortByTout(hospital):
    return hospital["tout_required"]

def sortByToutActually(hospital):
    return hospital["Tout_actually"]

def sortByTeos(hospital):
    return hospital["eos_req"]

# def change_eos_for_tout(hospital_data, diff,hospitals_output):
#     # print(timedelta(minutes=diff))
#     Subtract_from_eos = timedelta(minutes=diff)
#     last_eos = hospital_data[0]["eos_req"]
#     update_eos = last_eos - Subtract_from_eos
#     hospital_data[0]["eos_req"] = update_eos
#     print(update_eos)
#
#     recursion_for_tout(hospital_data,hospitals_output)

def recursion_for_tout(hospital_data):
    min_to_add = 15
    add_to_teos = timedelta(minutes=min_to_add)
    first_teos = hospital_data[0]["eos_req"]
    for h in hospital_data:
        Tout_actually = first_teos + add_to_teos
        if Tout_actually > h["tout_required"]:
                diff = Tout_actually - h["tout_required"]
                diff_to_int = diff.seconds / 60
                Subtract_from_eos = timedelta(minutes=diff_to_int)
                last_eos = hospital_data[0]["eos_req"]
                updated_eos = last_eos - Subtract_from_eos
                hospital_data[0]["eos_req"] = updated_eos
                recursion_for_tout(hospital_data)

        h["Tout_actually"] = Tout_actually
        min_to_add += 5
        add_to_teos = timedelta(minutes=min_to_add)



def main_algorithm_calculation(batches,hospitals_output,batches_general_data):
    for b in batches: #calculate T1,Tout, Tcal,Teos, delivery_order, activity
        if not len(b)==0:
            index = batches.index(b)

            hospitals=[]

            hospital_data =[]
            for order in b:
                hospital_name=order['Name']
                if hospital_name not in hospitals: # order[1] = hospital name in order
                    hospitals.append(hospital_name)

                    #add hospital record to hospitals_output list
                    try:
                        hospital = next(h for h in hospitals_output if h["Name"] == order["Name"] and h["Batch"]==index+1)
                    except:
                        hospitals_output.append({"Name":order["Name"],"Activity":0,'Batch':index+1 })

                    if index != 1:  #condition for choosing Transport_time (min/max) - according to number of batch
                        tout_temp = order['injection_time'] - timedelta(minutes=(order['Transport_time_max'])) #T1-Transport_time
                        hospital_data.append({'hospital_name':hospital_name,'batch':index+1, "injection_time": order['injection_time'], "Transport_time": order['Transport_time_max'], "tout_required": tout_temp})

                    else:
                        tout_temp = order['injection_time'] - timedelta(minutes=(order['Transport_time_min'])) #T1-Transport_time
                        hospital_data.append({'hospital_name':hospital_name,'batch':index+1, "injection_time": order['injection_time'], "Transport_time": order['Transport_time_min'], "tout_required": tout_temp})
                    # print(hospital_name, " req: ", tout_temp)
            hospital_data.sort( key=sortByTout) #sort hospital_data by tout_temp

            Hospital_delivery_order=1
            interval_5 = 15
            minutes_for_eos_cal = timedelta(minutes=interval_5)
            for h in hospital_data:
                #delivery_order
                h["delivery_order"] = Hospital_delivery_order
                hospital_record = next(hospital for hospital in hospitals_output if
                                       hospital["Name"] == h['hospital_name'] and hospital['Batch'] == h['batch'])
                hospital_record['delivery_order'] = Hospital_delivery_order
                Hospital_delivery_order+=1

                eos_req = h["tout_required"] - minutes_for_eos_cal  # tout - intervals consider the order
                h["eos_req"] = (eos_req)
                interval_5 += 5
                minutes_for_eos_cal = timedelta(minutes=interval_5)
                # hospital_data.sort(key=sortBy)  # sort hospital_data by tout_eos

            hospital_data.sort(key=sortByTeos)  # sort hospital_data by tout_eos
            # print(hospital_data)
            recursion_for_tout(hospital_data)

            #save tout actual for each hospital
            for h in hospital_data:
                hospital_record = next(hospital for hospital in hospitals_output if
                                       hospital["Name"] == h['hospital_name'] and hospital['Batch'] == h['batch'])
                hospital_record['Tout_actually'] = h['Tout_actually']
                # print(h["hospital_name"] , " actually: ", h['Tout_actually'])

            # insert Teos - new module
            Teos_key = "Teos"
            t_eos_final = hospital_data[0]["eos_req"] #first index in the list is the shortest time of eos (because it's sorted)
            batches_general_data[index][Teos_key] = t_eos_final

            # insert Tout
            Tout_key = "Tout"
            hospital_data.sort(key=sortByToutActually)
            last_index = len(hospital_data)-1
            t_out_final = hospital_data[last_index]['Tout_actually'] #last tout actually
            batches_general_data[index][Tout_key] = t_out_final

            #insert Tcal - new module
            Tcal_key = "Tcal"
            minutes_for_Tcal_cal = timedelta(minutes=30)
            t_cal = t_out_final + minutes_for_Tcal_cal
            batches_general_data[index][Tcal_key] = t_cal


            bottles_b = len(hospitals)
            bottels_key = "bottles_mum"
            batches_general_data[index][bottels_key] = bottles_b

            batches_general_data[index]["Activity"] = 0  # define the key
            for order in b:
                # insert (A) Activity_Tcal - Activity for Tcal time
                A_Tcal_key = "Activity_Tcal"
                injection_time= order["injection_time"]
                T_cal = batches_general_data[index]["Tcal"]
                A=order["Fixed_activity_level"]

                diff = (injection_time-T_cal).total_seconds() /60  #convert to minutes and then to float
                A_Tcal = math.ceil(A * math.pow((math.e),diff*lamda))  # math.ceil is round up


                if batches_general_data[index]["Activity"] + A_Tcal >= max_activity_batch:
                    batches[index+1].append(order)
                    batches[index].remove(order)
                    main_algorithm_calculation(batches,hospitals_output,batches_general_data)

                else:
                    # try:
                    #     hospital = next(h for h in hospitals_output if h["Name"] == order["Name"] and h["Batch"]==index+1)
                    # except:
                    #     hospitals_output.append({"Name":order["Name"],"Activity":0,'Batch':index+1 })
                    hospital = next(h for h in hospitals_output if h["Name"] == order["Name"] and h["Batch"] == index + 1)
                    order[A_Tcal_key] = A_Tcal
                    order['diff_Tcal_injectionT'] = diff
                    hospital["Activity"] +=A_Tcal

                    batches_general_data[index]["Activity"] += A_Tcal*1.05

def export_WP_Excel( selected_material, selected_date, all_batches_output, hospitals_output, batches_general_data):
    FilePath = "FDG format.xlsx"

    wb = load_workbook(FilePath)

    sheet = wb.active
    sheet = wb['work plan']

    hospitals = []
    row_index = 8

    for order in all_batches_output:
        if order['Name'] not in hospitals:
            hospitals.append(order['Name'])
            grey = "c0c0c0"
            col_start = 4
            col_end = 11

            sheet.cell(row=row_index, column=col_start).fill = PatternFill(start_color=grey, end_color=grey,
                                                                           fill_type="solid")  # bg of buffer cell

            merge_buffer = sheet.merge_cells(start_row=row_index, start_column=col_start, end_row=row_index,
                                             end_column=col_end)

            #hospital name in the first line
            col_s=12
            col_e=14
            sheet.cell(row=row_index, column=col_s).fill = PatternFill(start_color=grey, end_color=grey,
                                                                           fill_type="solid")  # bg of buffer cell
            sheet.cell(row=row_index, column=col_s).value =  order['Name']
            sheet.merge_cells(start_row=row_index, start_column=col_s,
                              end_row=row_index, end_column=col_e)

            #hospital name on the left side
            col_index = 3
            row_index += 1
            hospital_orders = [row for row in all_batches_output if row['Name'] == order['Name']]

            end_row_to_merge = row_index + len(hospital_orders) - 1

            hospital_name_cell = sheet.cell(row=row_index, column=col_index)
            hospital_name_cell.value = order['Name'] # insert hoapital name to the first col
            merge_hospital_name_cells = sheet.merge_cells(start_row=row_index, start_column=col_index,
                                                          end_row=end_row_to_merge, end_column=col_index)
            #sum activity of hospital for each batch
            hospital_output_data = [h for h in hospitals_output if h["Name"] == order["Name"] ]
            for h_b in hospital_output_data:
                if h_b["Batch"]==1:
                    sheet.cell(row=row_index, column=12).value = h_b['Activity']
                    sheet.cell(row=row_index, column=12).font = Font(size=60,bold=True)
                elif h_b["Batch"]==2:
                    sheet.cell(row=row_index, column=13).value = h_b['Activity']
                    sheet.cell(row=row_index, column=13).font = Font(size=60,bold=True)
                else:
                    sheet.cell(row=row_index, column=14).value = h_b['Activity']
                    sheet.cell(row=row_index, column=14).font = Font(size=60,bold=True)

            for row in hospital_orders:
                DosemCi = row['Fixed_activity_level']
                batch_num=row['batch']
                sheet.cell(row=row_index, column=4).value = row['DoseNumber']  # serial number
                sheet.cell(row=row_index, column=5).value = batch_num
                sheet.cell(row=row_index, column=6).value = DosemCi
                sheet.cell(row=row_index, column=7).value = batches_general_data[batch_num-1]['Teos']
                sheet.cell(row=row_index, column=8).value = batches_general_data[batch_num-1]['Tcal']
                sheet.cell(row=row_index, column=9).value = str(row['injection_time'])  # injection time
                sheet.cell(row=row_index, column=10).value = row['diff_Tcal_injectionT']
                sheet.cell(row=row_index, column=11).value = row['Activity_Tcal']

                row_index += 1

    batch_number = 0
    for b in batches_general_data:
        batch_number+=1
        if not len(b)==0:
            if batch_number==1:
                col_num = 12
            elif batch_number==2:
                col_num = 13
            else:
                col_num = 14

            sheet.cell(row=1, column=col_num).value = b['Tout']
            sheet.cell(row=3, column=col_num).value = b["bottles_mum"] #number of hospitals
            sheet.cell(row=4, column=col_num).value = b['Tcal']
            sheet.cell(row=5, column=col_num).value = b['Teos']
            sheet.cell(row=7, column=col_num).value = b['Activity']

    sheet2 = wb['more info']
    row_num=2
    b=0
    for hb in hospitals_output:
        if hb['Batch']==1:
            col_num = 1
            if not b==1:
                row_num=2
                b=1
        elif hb['Batch']==2:
            col_num = 4
            if not b==2:
                row_num=2
                b=2
        else:
            col_num = 6
            if not b==3:
                row_num=2
                b=3
        # print(hb['delivery_order'], " ",hb['Name'], " ",hb['Tout_actually'] )
        sheet2.cell(row=row_num, column=col_num).value =hb['delivery_order']
        sheet2.cell(row=row_num, column=col_num+1).value = hb['Name']
        sheet2.cell(row=row_num, column=col_num+2).value = hb['Tout_actually']
        row_num+=1

    downloads_path = str(Path.home() / "Downloads") + '/'

    wb_name = downloads_path + selected_material + 'workplan' + selected_date + '.xls'
    wb.save(wb_name)
    webbrowser.open(downloads_path)

def final_sort_by_hospital(batches):
    for b in batches:
        if not len(b)==0:
            b.sort(key=itemgetter('Name'))
    return batches

def flat_list(batches):
    flat_list = []
    batch_num=0
    for b in batches:
        batch_num+=1
        if not len(b)==0:
            for order in b:
                order['batch'] = batch_num
                flat_list.append(order)
    return flat_list
#end algorithm functions


class Popup(Toplevel):
    def __init__(self):
        Toplevel.__init__(self)
        # self.popup = self

    def open_pop(self, title,geometry ):
        # self.geometry("900x550")
        self.geometry(geometry)
        self.title(title)
        Label(self, text=title, font=('Helvetica 17 bold'), fg='#034672').place( x=10, y=18)

        ## in line
        # #labels and entry box
        # p_last_label_x=20
        # p_last_label_y=80
        # i=0
        # column_num=1
        #
        # for lab in labels:
        #     p_label = Label(self, text=lab[0])
        #     p_label.grid(row=1, column=column_num)
        #     p_label.place(x=p_last_label_x, y=p_last_label_y)
        #
        #     # Entry boxes
        #     entry_box = Entry(self, width=15)
        #     entry_box.grid(row=2, column=column_num)
        #     entry_box.place(x=p_last_label_x + 3, y=p_last_label_y + 40)
        #
        #
        #     column_num+=1
        #     if lab[1]!= '':
        #         p_label_units = Label(self, text=lab[1])
        #         font = ("Courier", 9)
        #         p_label_units.config(font=("Courier", 9))
        #         p_label_units_x = p_last_label_x + p_label.winfo_reqwidth()-3
        #         p_label_units.place(x=p_label_units_x, y=p_last_label_y + 7)
        #
        #         if entry_box.winfo_reqwidth() > p_label.winfo_reqwidth()+p_label_units.winfo_reqwidth():
        #             p_last_label_x += entry_box.winfo_reqwidth() + 30
        #         else:
        #             p_last_label_x += p_label.winfo_reqwidth()  +p_label_units.winfo_reqwidth()+ 30
        #     else:
        #         p_last_label_x += entry_box.winfo_reqwidth() + 30



    def is_legal(self, table_name, entries, error_labels_list):
        #validation-  not null filed is not empty
        column_input = dict_input_column[table_name]
        datatype_notnull_column = NOT_NULL_DataType_col(table_name)
        datatype_in_db = [data[1:] for data in datatype_notnull_column if data[0] == table_name and data[1] in column_input]
        input_values_list = self.get_entry(entries)

        for error_lab in error_labels_list: #inite error labeles (for more than one tries)
            error_lab['text'] = ""

        legal_notnull=True
        legal_datatype=True
        legal = True

        for col in datatype_in_db:
            if col[0] in column_input:
                i = column_input.index(col[0])   #index in input_values_list
                if col[2]=='NO' and input_values_list[i] == "":    # Not null validation
                    entries[i].config(bg=red_color)
                    error_labels_list[i]['text'] = "Please fill the box"
                    legal_notnull = False
                    legal=False

                else:  # data type validation

                    if col[1] == 'varchar':
                        try:
                            str(input_values_list[i])
                        except:
                            legal_datatype = False
                            entries[i].config(bg=red_color)
                            error_labels_list[i]['text'] = "Incorrect data format"
                    if col[1] == 'int':
                        try:
                            int(input_values_list[i])
                        except:
                            legal_datatype = False
                            entries[i].config(bg=red_color)
                            error_labels_list[i]['text'] = "Incorrect data format"
                    if col[1] == 'float':
                        try:
                            float(input_values_list[i])
                        except:
                            legal_datatype = False
                            entries[i].config(bg=red_color)
                            error_labels_list[i]['text'] = "Incorrect data format"

                    if col[1] == 'boolean':
                        try:
                            bool(input_values_list[i])
                        except:
                            legal_datatype = False
                            entries[i].config(bg=red_color)
                            error_labels_list[i]['text'] = "Incorrect data format"


                    if col[1] == 'time':
                        try:
                            datetime.strptime(input_values_list[i], '%H:%M').time()

                        except:
                            legal_datatype = False
                            entries[i].config(bg=red_color)
                            error_labels_list[i]['text'] = "Incorrect data format"

                    if col[1] == 'date':
                        try:
                            datetime.strptime(input_values_list[i], '%m/%d/%Y').date() or datetime.strptime(input_values_list[i], '%m-%d-%Y').date()
                        except:
                            legal_datatype = False
                            entries[i].config(bg=red_color)
                            error_labels_list[i]['text'] = "Incorrect data format"


        if not legal_notnull:
            text = "There are unallowed empty box. Please fill the highlighted fiels"
            error_message(text)
            self.lift()

        if not legal_datatype:
            legal=False
            error_message("Incorrect data format in highlighted box")
            self.lift()

         #move popup to front
        return legal


    def update_record(self,query, pk,list, update_values_list):
        selected = list.focus()
        #show the changes
        list.item(selected, text="", values = update_values_list)

        #save new values in the db
        updateCyclotronInDB = query
        try:
            cursor.execute(updateCyclotronInDB, update_values_list)
            db.commit()
        except:
            # Rollback in case there is any error
            db.rollback()

        self.destroy()


    def cancel_popup(self):
        self.destroy()


    def save_cancel_button(self, save_title,on_click_save_fun, *args):
        save_button = Button(self, text=save_title,
                               command=lambda: on_click_save_fun(*args))

        save_button.pack(side=LEFT)
        save_button_position_x = self.winfo_screenheight() / 2 - save_button.winfo_reqwidth()/2 +20
        save_button_position_y = 485
        # save_button_position_y = self.winfo_screenheight() *0.6 - save_button.winfo_reqheight()/2


        save_button.place(x=save_button_position_x, y=save_button_position_y)

        cancel_button = Button(self, text="Cancel", command=lambda: self.cancel_popup())
        cancel_button.pack(side=LEFT)
        cancel_button.place(x=save_button.winfo_reqwidth() + save_button_position_x + 10, y=save_button_position_y)


    def update_if_selected(self,query,pk,list,table_name,entries,error_labels_list):
        update_values_list=self.get_entry(entries)
        update_values_list.append(pk)
        if update_values_list is None: #if the user dont select record
            error_message("Please select record")
        else:
            legal = self.is_legal(table_name, entries,error_labels_list )
            if legal:
                self.update_record(query, pk,list,update_values_list)
            # else:
            #     text = "There are unallowed empty box. Please fill the empty fiels"
            #     error_message(text)

                self.destroy()


    def get_entry(self, entries): # to edit_popup - get user changes in entry box
        update_values_list=[]

        for entry in entries:
            entry.config(bg='white')
            update_values_list.append(entry.get())
        return update_values_list

    def edit_popup(self, labels, valueList, save_title, *args):
        # labels and entry box
        p_last_label_x = 30
        p_last_label_y = 80
        value_index = 0
        row_num = 1

        # grab record values

        # prevented 'Date', 'Batch Number','Material' show as entry box
        if args[len(args) - 1] == 'batch' :
            label_text = valueList [0] + '  |  '+ valueList[1]  +'  |  Batch Number: '+ valueList[2]
            p_label = Label(self, text=label_text, font=('Helvetica 14 bold '), fg=label_color)
            p_label.grid(row=row_num, column=1)
            p_last_label_y-=18
            p_label.place(x=p_last_label_x, y=p_last_label_y)
            valueList=valueList[3:]
            p_last_label_y += 33
            p_last_label_x+=10


        entries = []
        error_labels_list=[]
        for lab in labels:
            p_label = Label(self, text=lab[0])
            p_label.grid(row=row_num, column=1)
            p_label.place(x=p_last_label_x, y=p_last_label_y)

            row_num += 1

            # Entry boxes
            entry_box = Entry(self, width=20)
            entry_box.grid(row=row_num, column=2)
            entry_box.place(x=p_last_label_x + 4, y=p_last_label_y + 30)

            # insert value into entry box
            entry_box.insert(0, valueList[value_index])
            value_index += 1
            entries.append(entry_box)

            if args[len(args) - 1] == 'batch' and lab[0] in ('Time leaves Hadassah','Total EOS','EOS Time'):
                entry_box.config(state='disabled')
                p_last_label_y+=entry_box.winfo_reqheight()

            if lab[1] != '':
                p_label_units = Label(self, text=lab[1])
                font = ("Courier", 9)
                p_label_units.config(font=("Courier", 9))
                p_label_units_x = p_last_label_x + p_label.winfo_reqwidth()
                p_label_units.place(x=p_label_units_x, y=p_last_label_y + 5)


            # p_last_label_y += entry_box.winfo_reqheight() + 35 + p_label.winfo_reqheight()
            # row_num += 1

            p_last_label_y += entry_box.winfo_reqheight() + p_label.winfo_reqheight()
            if args[len(args) - 1] == 'batch' and lab[0] in ('Time leaves Hadassah','Total EOS','EOS Time'):
                pass
            else:
                # error labels
                error_label = Label(self, text='', font=('Courier', 8), fg='red')
                error_label.place(x=p_last_label_x + 1, y=p_last_label_y+6)
                error_labels_list.append(error_label)

                p_last_label_y += 18 + error_label.winfo_reqheight()
            row_num += 1

        self.save_cancel_button(save_title, self.update_if_selected, *args, entries,error_labels_list)

    def Add_if_legal(self, Addquery, list,table_name, entries, error_labels_list):
        legal = self.is_legal(table_name, entries,error_labels_list)
        if legal:
            input_values_list = self.get_entry(entries)
            try:
                #insert the record to db
                cursor.execute(Addquery, input_values_list)
                db.commit()

                #insert the id from db to values list (not in table) to allow deleting the record without refreshing the page
                pk_name = [pk[1] for pk in table_pk_list if pk[0] == table_name][0]
                selectMaxIDquery2 = "SELECT MAX(" + pk_name + ") FROM " + table_name
                cursor.execute(selectMaxIDquery2)
                data = cursor.fetchall()
                input_values_list.append(data[0][0])

                list.insert(parent='', index='end', iid=None, text='',
                            values=input_values_list)

            except:
                # Rollback in case there is any error
                db.rollback()

            self.destroy()

    # def export_WP_To_Excel(self,selected_date, selected_material, data):
    #     # ordersQuery = """SELECT h.Name,h.Fixed_activity_level , o.injection_time,o.amount, m.materialName, o.Date
    #     #                             FROM hospital h INNER JOIN orders o ON  h.idhospital=o.hospitalID INNER JOIN material m ON m.idmaterial=o.materialID
    #     #                             where Date = '""" + selected_date + """' and m.materialName= '""" + selected_material + "' ORDER BY hospitalID, injection_time "
    #     #
    #     # cursor.execute(ordersQuery)
    #     # data = cursor.fetchall()
    #
    #     FilePath = "FDG format.xlsx"
    #
    #     wb = load_workbook(FilePath)
    #
    #     sheet = wb.active
    #     sheet = wb['work plan']
    #
    #     hospitals = []
    #     row_index = 9
    #     for order in data:
    #         if order[0] not in hospitals:
    #             grey = "c0c0c0"
    #             col_start = 4
    #             col_end = 16
    #
    #             sheet.cell(row=row_index, column=col_start).fill = PatternFill(start_color=grey, end_color=grey,
    #                                                                            fill_type="solid")  # bg of buffer cell
    #
    #             merge_buffer = sheet.merge_cells(start_row=row_index, start_column=col_start, end_row=row_index,
    #                                              end_column=col_end)
    #
    #             i = 1
    #             col_index = 3
    #             row_index += 1
    #             hospital_orders = [row[1:] for row in data if row[0] == order[0]]
    #             end_row_to_merge = row_index + len(hospital_orders) - 1
    #             hospital_name_cell = sheet.cell(row=row_index, column=col_index)
    #             hospital_name_cell.value = order[0]  # insert hoapital name to the first col
    #             merge_hospital_name_cells = sheet.merge_cells(start_row=row_index, start_column=col_index,
    #                                                           end_row=end_row_to_merge, end_column=col_index)
    #             hospitals.append(order[0])
    #
    #             for row in hospital_orders:
    #                 DosemCi = row[0] * row[2]
    #                 # sheet.cell(row=row_index, column=4).value = i  # serial number
    #                 # sheet.cell(row=row_index, column=6).value = DosemCi
    #                 # sheet.cell(row=row_index, column=11).value = str(row[1])  # injection time
    #
    #                 sheet.cell(row=row_index, column=4).value = i  # serial number
    #                 sheet.cell(row=row_index, column=6).value = DosemCi
    #                 sheet.cell(row=row_index, column=9).value = str(row[1])  # injection time
    #                 i += 1
    #                 row_index += 1
    #                 downloads_path = str(Path.home() / "Downloads")
    #     downloads_path = str(Path.home() / "Downloads") +'/'
    #
    #     wb_name = downloads_path+selected_material+ 'workplan'+ selected_date +'.xls'
    #     wb.save(wb_name)
    #     webbrowser.open(downloads_path)


    def legal_wp(self,selected_material,selected_date,error_labels_list,selected_material_ID,dataLen ):
        legal = True

        exist_wp_query = "SELECT * FROM workplan WHERE Date= '" +selected_date+ """'
                            AND ISNULL(deleted) AND materialID=""" + str(selected_material_ID)
        cursor.execute(exist_wp_query)
        exist_wp_data = cursor.fetchall()

        for error_lab in error_labels_list: #inite error labeles (for more than one tries)
            error_lab['text'] = ""

        if selected_material=='Select a material':
            error_message('Please select a material')
            # entries[0].config(bg=red_color)
            error_labels_list[0]['text'] = "Please select a material"
            legal = False

        elif len(exist_wp_data)!=0:
            error_text = "There is a work plan for " + selected_material + " for date " + selected_date + " in the system. Identical work plans cannot be created."
            error_message(error_text)
            self.lift()
            return False

        else:
            if dataLen == 0:
                error_text = "There are no orders for material " + selected_material + " for date " + selected_date + " in the system. Please change your selection"
                error_message(error_text)
                self.lift()
                legal = False

        if not legal:
            self.lift()
        return legal


    def select_resources(self, selected_date, selected_material, data):
        # labels and entry box
        p_last_label_x = 30
        p_last_label_y = 80
        value_index = 0
        row_num = 1
        labels = ['Cyclotron', 'Module', 'Module', 'Module']
        entries = []
        error_labels_list = []
        rec_var_list=[]

        for lab in labels:
            p_label = Label(self, text=lab)
            p_label.grid(row=row_num, column=1)
            p_label.place(x=p_last_label_x, y=p_last_label_y)
            row_num += 1

            if lab == 'Cyclotron':
                rec_var = StringVar(self)
                rec_var.set("Select a Cyclotron")  # default value

                query = "SELECT version,idresourceCyclotron FROM resourcecyclotron"
                cursor.execute(query)
                rec_options_list = cursor.fetchall()


            elif lab == 'Module':
                rec_var = StringVar(self)
                rec_var.set("Select a module")  # default value

                query = "SELECT version,idresourcemodule FROM resourcemodule"
                cursor.execute(query)
                rec_options_list = cursor.fetchall()

            rec_var_list.append(rec_var)
            recname = [m[0] for m in rec_options_list]
            rec_dropdown = OptionMenu(self, rec_var, *recname)
            rec_dropdown.place(x=p_last_label_x + 4, y=p_last_label_y + 30)
            p_last_label_y += rec_dropdown.winfo_reqheight() + p_label.winfo_reqheight()

            # error labels
            error_label = Label(self, text='', font=('Courier', 8), fg='red')
            error_label.place(x=p_last_label_x + 1, y=p_last_label_y)
            error_labels_list.append(error_label)
            row_num += 1

            p_last_label_y += 15 + error_label.winfo_reqheight()

            # buttons
            save_button = Button(self, text='Create a work plan',
                                 command=lambda: self.create_wp_popup(rec_var_list,selected_date, selected_material, data,error_labels_list,OptionMenu))

            save_button.pack(side=LEFT)
            save_button_position_x = self.winfo_screenheight() / 2 - save_button.winfo_reqwidth() / 2
            #
            save_button_position_y = self.winfo_screenwidth() / 2 - save_button.winfo_reqwidth()

            save_button.place(x=save_button_position_x, y=save_button_position_y)

            cancel_button = Button(self, text="Cancle", command=lambda: self.cancel_popup())
            cancel_button.pack(side=LEFT)
            cancel_button.place(x=save_button.winfo_reqwidth() + save_button_position_x + 10, y=save_button_position_y)



    def create_wp_popup(self,  selected_date, selected_material):
        # algorithm
        query = """SELECT o.idorders, h.Name,o.DoseNumber,h.Fixed_activity_level*o.amount as Fixed_activity_level, o.injection_time,o.amount,h.Transport_time_min,h.Transport_time_max
                FROM hospital h INNER JOIN orders o ON  h.idhospital=o.hospitalID INNER JOIN material m ON m.idmaterial=o.materialID
                where Date = '""" + str(selected_date) + """ ' and m.materialName= '""" +str(selected_material)+ """' ORDER BY injection_time """

        cursor = db.cursor(dictionary=True)
        cursor.execute(query)
        data = cursor.fetchall()
        print('date: ', data)
        cursor = db.cursor(dictionary=False)

        batch1 = []
        batch2 = []
        batch3 = []
        batch3_exist = True

        for order in data:
            order_time = datetime.strptime(str(order["injection_time"]), '%H:%M:%S').time()
            if order_time < datetime.strptime('15:00:00', '%H:%M:%S').time():  # batch 1
                batch1.append(order)

            elif order_time < datetime.strptime('23:00:00', '%H:%M:%S').time():  # batch 2
                batch2.append(order)
            else:  # batch 3
                batch3.append(order)

        dict_batch1_general = {}
        dict_batch2_general = {}
        dict_batch3_general = {}
        batches_general_data = [dict_batch1_general, dict_batch2_general, dict_batch3_general]
        batches = [batch1, batch2, batch3]


        hospitals_output = []  # for output
        main_algorithm_calculation(batches, hospitals_output, batches_general_data)
        hospitals_output.sort(key=lambda hb:(hb['Batch'],hb['delivery_order']))
        print("batches: ",batches)
        print("hospitals_output: ",hospitals_output)
        print("batches_general_data: ",batches_general_data)
        all_batches_output = flat_list(batches)
        all_batches_output.sort(key=itemgetter('Name'))
        cursor = db.cursor()

        selected_material_ID = next(m[1] for m in material_options_list if m[0]==selected_material)

        #create wp record
        new_wp_list =(str(selected_date),selected_material_ID)
        work_plan_query = "INSERT INTO workplan (Date, materialID) VALUES " + str(new_wp_list)
        cursor.execute(work_plan_query)
        db.commit()

        #get workplanID (for batch records)
        workplanID_Query = "SELECT MAX(idworkplan) FROM workplan "
        cursor.execute(workplanID_Query)
        workplanID_list = cursor.fetchall()
        workplanID = workplanID_list[0][0]

        batch_input_values_list=[]
        index=1
        for batch in batches_general_data:
            if len(batch)!=0:
                values= (workplanID,index,str(batch['Tout']), batch['Activity'] ,str(batch['Teos']))
                create_batch_query="INSERT INTO batch (workplanID, batchNumber, Time_leaves_Hadassah,Total_eos,EOS_TIME) VALUES " + str(values)

                cursor.execute(create_batch_query)
                db.commit()

                #for ui table - in batch page
                list = [str(selected_date),selected_material,index,str(batch['Tout']),batch['Activity'],batch['Teos'],None,None]
                batch_input_values_list.append(list)
            index+=1

        i=1
        for b in batches:
            if len(b)!=0:
                # get batchID (for orders records)
                batchID_Query = "SELECT idbatch FROM batch WHERE workplanID= " + str(workplanID) + """
                                AND batchNumber = """ + str(i)
                cursor.execute(batchID_Query)
                batchID_list = cursor.fetchall()
                batchID = batchID_list[0][0]

                #for ui table - in batch page
                batch_input_values_list[i-1].append(batchID)

                for order in b:
                    values= (batchID,float(order['Activity_Tcal']))
                    update_rec_query = """UPDATE orders SET batchID= %s,DecayCorrected= %s
                                         WHERE idorders = """ + str(order['idorders'])

                    cursor.execute(update_rec_query, values)
                    db.commit()
            i+=1

        # insert the id from db to values list (not in table) to allow deleting the record without refreshing the page
        wp_input_values_list=[str(selected_date),selected_material]
        selectMaxIDquery = """SELECT MAX(idworkplan) FROM workplan"""
        cursor.execute(selectMaxIDquery)
        data = cursor.fetchall()
        wp_input_values_list.append(data[0][0])

        #add to wp table (show to user)
        wp_tabel.insert(parent='', index='end', iid=None, text='',
                    values=wp_input_values_list)


        # add to batch table (show to user)
        for b_r in batch_input_values_list:
            batch_tabel.insert(parent='', index='end', iid=None, text='',
                           values=b_r)

        #excel
        excelIcon = Image.open("excelIcon.png")
        resizedExcelIcon = excelIcon.resize((40, 40), Image.ANTIALIAS)
        imgExcel = ImageTk.PhotoImage(resizedExcelIcon)
        # ExcelButton = Button(self, image=imgExcel, borderwidth=0,
        #                      command=lambda: self.export_WP_To_Excel(selected_date, selected_material, data))
        ExcelButton = Button(self, image=imgExcel, borderwidth=0,
                             command=lambda: export_WP_Excel(selected_material,selected_date,all_batches_output,hospitals_output,batches_general_data))
        # ExcelButton.pack(side=LEFT)
        ExcelButton.place(x=90, y=90 )

        Label(self, text='Export to Excel File', font=('Helvetica 12'), fg='grey').place(
            x=70 - ExcelButton.winfo_reqwidth() / 2, y=90 + ExcelButton.winfo_reqheight())

        root.mainloop()


    def wp_validation_plus(self,material_var,cal,error_labels_list):

        selected_date = cal.get()
        selected_material = material_var.get()
        cursor = db.cursor()

        selected_material_ID = next(m[1] for m in material_options_list if m[0]==selected_material)

        ordersQuery = """SELECT h.Name,h.Fixed_activity_level , o.injection_time,o.amount, m.materialName, o.Date
                                                          FROM hospital h INNER JOIN orders o ON  h.idhospital=o.hospitalID INNER JOIN material m ON m.idmaterial=o.materialID
                                                          where Date = '""" + selected_date + """' and o.materialID= '""" + str(selected_material_ID) + "' ORDER BY hospitalID, injection_time "

        cursor.execute(ordersQuery)
        data = cursor.fetchall()
        popup_size= "850x550"
        dataLen=len(data)

        legal = self.legal_wp(selected_material,selected_date,error_labels_list, selected_material_ID,dataLen)
        if legal:
            self.destroy()
            # export_popup=Popup()
            # title = 'Work Plan - '+selected_material +' '+ selected_date
            # export_popup.open_pop(title,popup_size)
            # export_popup.wp_popup(selected_date, selected_material,data)

            select_rec = Popup()
            title = 'Work Plan - ' + selected_material + ' ' + selected_date
            select_rec.open_pop(title, popup_size)
            # select_rec.select_resources(selected_date, selected_material, data)

            select_rec.create_wp_popup(selected_date, selected_material)


    def add_wp_popup(self):
        # labels and entry box
        p_last_label_x = 30
        p_last_label_y = 80
        value_index = 0
        row_num = 1
        labels = ['Material','Date']
        entries=[]
        error_labels_list=[]

        for lab in labels:
            p_label = Label(self, text=lab)
            p_label.grid(row=row_num, column=1)
            p_label.place(x=p_last_label_x, y=p_last_label_y)
            row_num += 1

            if lab == 'Material':
                material_var = StringVar(self)
                material_var.set("Select a material")  # default value



                materialname = [m[0] for m in material_options_list ]
                material_dropdown = OptionMenu(self, material_var, *materialname)
                material_dropdown.place(x=p_last_label_x + 4, y=p_last_label_y + 30)
                p_last_label_y += material_dropdown.winfo_reqheight() + p_label.winfo_reqheight()

            elif lab=='Date':
                # Add Calendar
                cal = DateEntry(self, width=12, background='darkblue',
                                foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
                cal.place(x=p_last_label_x + 4, y=p_last_label_y + 30)
                p_last_label_y += cal.winfo_reqheight() + p_label.winfo_reqheight()

            # error labels
            error_label = Label(self, text='', font=('Courier', 8), fg='red')
            error_label.place(x=p_last_label_x + 1, y=p_last_label_y+1)
            error_labels_list.append(error_label)
            row_num += 1

            p_last_label_y += 18 + error_label.winfo_reqheight()

            #buttons
            next_button = Button(self, text='Create work plan',
                                 command=lambda: self.wp_validation_plus( material_var,cal,error_labels_list ))

            next_button.pack(side=LEFT)
            save_button_position_x = self.winfo_screenheight() / 2 - next_button.winfo_reqwidth() / 2
            #
            save_button_position_y = self.winfo_screenheight() / 2

            next_button.place(x=save_button_position_x, y=save_button_position_y)

            cancel_button = Button(self, text="Cancel", command=lambda: self.cancel_popup())
            cancel_button.pack(side=LEFT)
            cancel_button.place(x=next_button.winfo_reqwidth() + save_button_position_x + 10, y=save_button_position_y)

            # #buttons
            # save_button = Button(self, text='Create a work plan',
            #                      command=lambda: self.create_wp(material_var,cal,error_labels_list ))
            #
            # save_button.pack(side=LEFT)
            # save_button_position_x = self.winfo_screenheight() / 2 - save_button.winfo_reqwidth() / 2
            # #
            # save_button_position_y = self.winfo_screenheight() / 2
            #
            # save_button.place(x=save_button_position_x, y=save_button_position_y)
            #
            # cancel_button = Button(self, text="Cancle", command=lambda: self.cancel_popup())
            # cancel_button.pack(side=LEFT)
            # cancel_button.place(x=save_button.winfo_reqwidth() + save_button_position_x + 10, y=save_button_position_y)


    def add_popup(self, labels, save_title, *args):
        # labels and entry box
        p_last_label_x = 30
        p_last_label_y = 80
        value_index=0
        row_num = 1

        # grab record values
        error_labels_list=[]
        entries = []
        for lab in labels:
            p_label = Label(self, text=lab[0])
            p_label.grid(row=row_num, column=1)
            p_label.place(x=p_last_label_x, y=p_last_label_y)
            row_num += 1

            # Entry boxes
            entry_box = Entry(self, width=20,insertbackground=label_color)
            entry_box.grid(row=row_num, column=2)
            entry_box.place(x=p_last_label_x + 4, y=p_last_label_y + 30)
            entries.append( entry_box)

            if lab[1] != '':
                p_label_units = Label(self, text=lab[1])
                font = ("Courier", 9)
                p_label_units.config(font=("Courier", 9))
                p_label_units_x = p_last_label_x + p_label.winfo_reqwidth()
                p_label_units.place(x=p_label_units_x, y=p_last_label_y + 5)

            p_last_label_y += entry_box.winfo_reqheight()  + p_label.winfo_reqheight()

            #error labels
            error_label = Label(self, text='', font=('Courier',8),fg='red' )
            error_label.place(x=p_last_label_x+1, y=p_last_label_y+6)
            error_labels_list.append(error_label)
            row_num += 1


            p_last_label_y += 18  + error_label.winfo_reqheight()
        self.save_cancel_button(save_title, self.Add_if_legal,*args, entries,error_labels_list ) # will add save.cancel buttons (and click on functions)


class table(ttk.Treeview):
    def  __init__(self,frame,scroll_width,list_height,side,x_crol,y_crol,lable_place_x,
                               lable_place_y):
        self.side = side
        scroll = Scrollbar(frame, orient="vertical", width=scroll_width)
        scroll.pack(side=side)
        scroll.place(x=x_crol, y=y_crol)

        ttk.Treeview.__init__(self,frame, yscrollcommand=scroll.set, height=list_height)
        self.pack(side=LEFT, padx=lable_place_x + 30, pady=lable_place_y + 50)
        scroll.config(command=self.yview )

        # list = self.(frame, yscrollcommand=scroll.set, height=list_height)

    # def create_fully_tabel(self,scroll_width,side, x_crol,y_crol, frame, list_height, lable_place_x,lable_place_y, columns_name_list, query):
    def create_fully_tabel(self, columns_name_list, query):

        # scroll = Scrollbar(frame, orient="vertical", width=scroll_width)
        # scroll.pack(side=side)
        # scroll.place(x=x_crol, y=y_crol)
        #
        # list = ttk.Treeview(frame, yscrollcommand=scroll.set, height=list_height)

        self['columns'] = columns_name_list

        self.column("#0", width=0, stretch=NO)
        self.heading("#0", text="", anchor=CENTER)

        i=0
        len_of_col=len(columns_name_list)
        for column_name in columns_name_list:
            # column format
            if i == 0 or i == len_of_col-2:
                width = len(column_name)*6 +30
            else:
                width = len(column_name)*6

            self.column(column_name, anchor=CENTER, width=width)
            # # Create Headings
            self.heading(column_name, text=column_name, anchor=CENTER)
        # query = query + " WHERE ISNULL(deleted)"
        cursor.execute(query)
        data = cursor.fetchall()
        # print(query,data)
        iid=0
        for recorf in data:
            val=[]
            for i in range (0,len_of_col): # plus 1 is for the pk that will not show in the table
                val.append(recorf[i+1])
            val.append(recorf[0])

            self.insert(parent='', index='end', iid=iid, text='',
                       values=val)
            iid +=1

            self.pack()

    def selected(self):
        selected = self.focus()
        selected_record = self.item(selected, 'values')
        return selected_record

    def selected_is_non(self, selected_record):
        if selected_record =='':
            text = "Please select a record from the table"
            error_message(text)
            return True
        else:
            return False


    def fk_rec_is_exist(self,query,table_name, pk_delected_record ):
        fk_list = [rec for rec in fk if rec[2]== table_name]
        if len(fk_list) != 0:
            for fk_rec in fk_list:
                query = "select * from "+ fk_rec[0] + " where " + fk_rec[1] + "=" + pk_delected_record
                try:
                    cursor.execute(query)
                    data = cursor.fetchall()
                    if data != []:
                        return True
                except:
                    #Rollback in case there is any error
                    db.rollback()
        return False

    def delete_record(self, query,table_name):
        selected_rec = self.selected()
        item_in_string= ', '.join([ item for item in selected_rec[:selected_rec.__len__()-1]])
        is_non=self.selected_is_non(selected_rec)
        if not is_non:
            len = selected_rec.__len__()
            pk_delected_record = selected_rec[len-1]
            title_tab = "Delete Record"
            text_mess= "Are you sure you want to delete " + item_in_string + " ?"
            if YES_NO_message(title_tab, text_mess):
                pk_delected_record_list = (pk_delected_record, )
                to_delete= not self.fk_rec_is_exist(query,table_name,pk_delected_record)
                pk_name = [pk[1] for pk in table_pk_list if pk[0] == table_name][0]
                if to_delete:
                    try:
                        query = "delete from "+table_name + " where " + pk_name + "=" + pk_delected_record
                        cursor.execute(query )
                        db.commit()
                    except:
                        # Rollback in case there is any error
                        db.rollback()

                else: #to hide
                    query2 = "UPDATE " + table_name +" SET deleted = True " +"WHERE " + pk_name + "=" + pk_delected_record
                    cursor.execute(query2)
                    db.commit()
                self.delete(self.selection()[0])


    def delete_WP_record(self):
        selected_rec = self.selected()
        item_in_string= ', '.join([ item for item in selected_rec[:selected_rec.__len__()-1]])
        is_non=self.selected_is_non(selected_rec)
        if not is_non:
            len = selected_rec.__len__()
            pk_delected_record = selected_rec[len-1]
            title_tab = "Delete Record"
            text_mess= "Are you sure you want to delete " + item_in_string + " ?"
            if YES_NO_message(title_tab, text_mess):
                try:
                    # hide work plan
                    hide_wp_query = "UPDATE workplan SET deleted = True WHERE idworkplan=" +  pk_delected_record
                    cursor.execute(hide_wp_query)
                    db.commit()
                except:
                    raise TypeError("update work plan error")

                try:
                    # update orders
                    update_orders_query = """UPDATE orders SET batchID=NULL,DecayCorrected=NULL WHERE batchID  IN
                                          (SELECT idbatch FROM batch WHERE workplanID=""" + pk_delected_record +")"
                    cursor.execute(update_orders_query)
                    db.commit()
                except:
                    raise TypeError("update orders error")

                try:
                    #for deleting rows from table
                    delete_table_batch_query = "SELECT idbatch FROM batch WHERE workplanID=" + pk_delected_record
                    cursor.execute(delete_table_batch_query)
                    to_delete_list_idbatch = cursor.fetchall()
                    to_delete_batches_list =  [b[0] for b in to_delete_list_idbatch]

                    # delete batch
                    delete_batch_query = "DELETE FROM batch WHERE workplanID=" + pk_delected_record
                    cursor.execute(delete_batch_query)
                    db.commit()
                except:
                    raise TypeError("update batch error")

                self.delete(self.selection()[0])

                # delete from batch table (show to user)
                children = batch_tabel.get_children()
                batch_table_index = [b for b in children if batch_tabel.item(b)['values'][7] in to_delete_batches_list]

                for b_index in batch_table_index:
                    batch_tabel.delete(b_index)

##################### settings - cyclotron #####################
#cyclotron frame
cycloSettingsFrame = Frame(root)
# h = Scrollbar(cycloSettingsFrame, orient='horizontal')
# cycloSettingsFrame.pack(fill=X)

# feed label - cyclotron
feedLabel = Label(cycloSettingsFrame, text = 'Settings ➝ ', font=label_font_flag,fg=label_color)
PlaceLable_X=50
PlaceLable_Y=10
feedLabel.pack(side=LEFT)
feedLabel.place(x=PlaceLable_X,y=PlaceLable_Y)

feedLabeflag = Label(cycloSettingsFrame, text = 'Cyclotron', font=label_font_flag_on_page,fg=label_color)

PlaceLable2_X=135
feedLabeflag.pack(side=LEFT)
feedLabeflag.place(x=PlaceLable2_X,y=PlaceLable_Y)

##################### Cyclotron #####################
# Cyclotron Details label
CyclotronLabel = Label(cycloSettingsFrame, text = 'Cyclotron Details', font=sub_label_font,fg=label_color)
Lable_place_x=80
Lable_place_y=60

CyclotronLabel.pack(side=LEFT)
CyclotronLabel.place(x=Lable_place_x,y=Lable_place_y)

###cycortion tabel###
scroll_width=20
tab_side=LEFT
x=613
y= 140
frame=cycloSettingsFrame
list_height=5
table_place_x = 80
table_place_y = 80
columns_name_list=('Version', 'Capacity (mci/h)', 'Constant Efficiency (mCi/mA)', 'Description')

query = "SELECT * FROM resourcecyclotron WHERE ISNULL(deleted)"

# cyclo_tabel=table(scroll_width,tab_side, x,y,frame,list_height,lable_place_x,lable_place_y, columns_name_list, query )
cyclo_tabel=table(frame,scroll_width,list_height,tab_side,x,y,table_place_x,
                  table_place_y,)
cyclo_tabel.create_fully_tabel( columns_name_list, query)


###cycortion functions###
def editCyclotronfun():
    selected_rec = cyclo_tabel.selected()
    selected_non=cyclo_tabel.selected_is_non(selected_rec)
    if not selected_non:
        editCyclPopup = Popup()
        popup_size = "900x550"
        editCyclPopup.open_pop('Edit Cyclotron Details',popup_size)

        query = "UPDATE resourcecyclotron SET version = %s ,capacity= %s, constant_efficiency= %s,description=%s  WHERE idresourceCyclotron = %s"
        pk = selected_rec[4]
        table_name = 'resourcecyclotron'
        labels = (('Version', ''), ('Capacity', '(mci/h)'), ('Constant Efficiency', '(mCi/mA)'), ('Description', ''))
        save_title = "Save Changes"

        editCyclPopup.edit_popup(labels, selected_rec, save_title, query, pk, cyclo_tabel,table_name)


def deleteCyclotronfun():
    query = "DELETE FROM resourcecyclotron WHERE idresourceCyclotron = %s"
    table_name='resourcecyclotron'
    cyclo_tabel.delete_record(query,table_name)

def addCyclotronfun():
    addCyclPopup = Popup()
    popup_size = "900x550"
    addCyclPopup.open_pop('Add Cyclotron Details',popup_size)
    labels = (('Version', ''), ('Capacity', '(mci/h)'), ('Constant Efficiency', '(mCi/mA)'), ('Description', ''))
    save_title = "Add Cyclotron"
    insertquery = "INSERT INTO resourcecyclotron SET version = %s ,capacity= %s, constant_efficiency= %s,description=%s"
    # selectIDquery = "SELECT MAX(idresourceCyclotron) FROM resourcecyclotron"
    table_name='resourcecyclotron'
    addCyclPopup.add_popup(labels, save_title, insertquery, cyclo_tabel,table_name)

#cyclotron buttons

#Create a button in the main Window to add record - cyclotron
cyclotronAddIcon = Image.open("addIcon.png")
resizedCycloAddIcon = cyclotronAddIcon.resize((25, 25), Image.ANTIALIAS)
imgAddCyclotron = ImageTk.PhotoImage(resizedCycloAddIcon)
addCyclotronButton = Button(cycloSettingsFrame, image=imgAddCyclotron, borderwidth=0, command=lambda : addCyclotronfun())
addCyclotronButton.pack(side= LEFT)
addCyclotronButton.place(x=table_place_x + cyclo_tabel.winfo_reqwidth() -100, y=table_place_y+14)

#Create a button in the main Window to edit  record (open the popup) - cyclotron
cyclotronEditIcon = Image.open("editIcon.jpg")
resizedCycloEditIcon = cyclotronEditIcon.resize((20, 20), Image.ANTIALIAS)
imgEditCyclotron = ImageTk.PhotoImage(resizedCycloEditIcon)
# editCyclotronButton = Button(ctcloSettingsFrame, image=imgEditCyclotron, borderwidth=0, command= lambda :editCyclotronfun())
editCyclotronButton = Button(cycloSettingsFrame, image=imgEditCyclotron, borderwidth=0, command= lambda :editCyclotronfun())

editCyclotronButton.pack(side= LEFT)
editCyclotronButton.place(x=table_place_x + cyclo_tabel.winfo_reqwidth() -50, y=table_place_y+15)


# Create a button in the main Window to Delete record - cyclotron
cyclotronDeleteIcon = Image.open("‏‏deleteIcon.png")
resizedCycloDeleteIcon = cyclotronDeleteIcon.resize((20, 20), Image.ANTIALIAS)
imgDeleteCyclotron = ImageTk.PhotoImage(resizedCycloDeleteIcon)
deleteCyclotronButton = Button(cycloSettingsFrame, image=imgDeleteCyclotron, borderwidth=0, command=lambda : deleteCyclotronfun())
deleteCyclotronButton.pack(side=LEFT)
deleteCyclotronButton.place(x=table_place_x + cyclo_tabel.winfo_reqwidth(), y=table_place_y + 15)


##################### settings - module #####################
#module frame
moduleSettingsFrame = Frame(root)
# h = Scrollbar(moduleSettingsFrame, orient='horizontal')
# moduleSettingsFrame.pack(fill=X)

# feed label - module
feedLabel = Label(moduleSettingsFrame, text = 'Settings ➝ ', font=label_font_flag,fg=label_color)
PlaceLable_X=50
PlaceLable_Y=10
feedLabel.pack(side=LEFT)
feedLabel.place(x=PlaceLable_X,y=PlaceLable_Y)

feedLabeflag = Label(moduleSettingsFrame, text = 'Module', font=label_font_flag_on_page,fg=label_color)

PlaceLable2_X=135
feedLabeflag.pack(side=LEFT)
feedLabeflag.place(x=PlaceLable2_X,y=PlaceLable_Y)

##################### Module #####################

# Module Details label
moduleLabel = Label(moduleSettingsFrame, text = 'Module Details', font=sub_label_font,fg=label_color)
# module_Lable_place_x=80
# module_Lable_place_y=60

moduleLabel.pack(side=LEFT)
moduleLabel.place(x=Lable_place_x,y=Lable_place_y)
moduleLabel.pack(side=RIGHT)
moduleLabel.place(x=Lable_place_x,y=Lable_place_y)

###module tabel###
scroll_width=20
tab_side=LEFT
x=420
y= 150
frame=moduleSettingsFrame
list_height=5
# table_place_x = 80
# table_place_y=80

columns_name_list=('Version', 'Capacity (mci/h)', 'Description')

queryModule = "SELECT * FROM resourcemodule WHERE ISNULL(deleted)"

module_tabel=table(frame,scroll_width,list_height,tab_side,x,y,table_place_x,
                   table_place_y)
module_tabel.create_fully_tabel( columns_name_list, queryModule)

###module functions###
def editModulefun():
    selected_rec = module_tabel.selected()
    selected_non = module_tabel.selected_is_non(selected_rec)
    if not selected_non:
        # popup_size = "800x450"
        popup_size = "900x550"
        editModulePopup = Popup()
        editModulePopup.open_pop('Edit Module Details', popup_size)

        query = "UPDATE resourcemodule SET version = %s ,capacity= %s, description=%s  WHERE idresourcemodule = %s"
        table_name = 'resourcemodule'
        pk = selected_rec[3]

        labels = (('Version', ''), ('Capacity', '(mci/h)'),  ('Description', ''))
        save_title = "Save Changes"

        editModulePopup.edit_popup(labels, selected_rec, save_title, query, pk, module_tabel,  table_name)


def addModulefun():
    addModulePopup = Popup()
    # popup_size = "800x450"
    popup_size = "900x550"
    addModulePopup.open_pop('Add Module Details',popup_size)
    labels = (('Version', ''), ('Capacity', '(mci/h)'), ('Description', ''))
    save_title = "Add Module"
    insetQuery = "INSERT INTO resourcemodule SET version = %s ,capacity= %s,description=%s"
    table_name='resourcemodule'
    addModulePopup.add_popup(labels, save_title, insetQuery, module_tabel, table_name)

def deleteModulefun():
    query = "DELETE FROM resourcemodule WHERE idresourcemodule = %s"
    table_name='resourcemodule'
    module_tabel.delete_record(query,table_name)


#module buttons

#Create a button in the main Window to add record - module
moduleAddIcon = Image.open("addIcon.png")
resizedModuleAddIcon = moduleAddIcon.resize((25, 25), Image.ANTIALIAS)
imgAddModule = ImageTk.PhotoImage(resizedModuleAddIcon)
addModuleButton = Button(moduleSettingsFrame, image=imgAddModule, borderwidth=0, command=addModulefun)
addModuleButton.pack(side= LEFT)
addModuleButton.place(x=table_place_x+ module_tabel.winfo_reqwidth() -100 , y=table_place_y+14)

#Create a button in the main Window to edit  record (open the popup) - module
moduleEditIcon = Image.open("editIcon.jpg")
resizedModuleEditIcon = moduleEditIcon.resize((20, 20), Image.ANTIALIAS)
imgEditModule = ImageTk.PhotoImage(resizedModuleEditIcon)
editModuleButton = Button(moduleSettingsFrame, image=imgEditModule, borderwidth=0, command=editModulefun)
editModuleButton.pack(side= LEFT)
editModuleButton.place(x=table_place_x+module_tabel.winfo_reqwidth() - 50, y=table_place_y+15)


#Create a button in the main Window to Delete record - module
moduleDeleteIcon = Image.open("‏‏deleteIcon.png")
resizedModuleDeleteIcon = moduleDeleteIcon.resize((20, 20), Image.ANTIALIAS)
imgDeleteModule = ImageTk.PhotoImage(resizedModuleDeleteIcon)
deleteModuleButton = Button(moduleSettingsFrame, image=imgDeleteModule, borderwidth=0, command=deleteModulefun)
deleteModuleButton.pack(side= LEFT)
deleteModuleButton.place(x=table_place_x+module_tabel.winfo_reqwidth(), y=table_place_y+15)



# ##################### Material #####################
##################### settings - Material #####################
#material frame
materialSettingsFrame = Frame(root)
# h = Scrollbar(materialSettingsFrame, orient='horizontal')
# materialSettingsFrame.pack(fill=X)

# feed label - material
feedLabelmaterial = Label(materialSettingsFrame, text = 'Settings ➝ ', font=label_font_flag,fg=label_color)
PlaceLable_X=50
PlaceLable_Y=10
feedLabelmaterial.pack(side=LEFT)
feedLabelmaterial.place(x=PlaceLable_X,y=PlaceLable_Y)

feedLabeflag = Label(materialSettingsFrame, text = 'Material', font=label_font_flag_on_page,fg=label_color)

PlaceLable2_X=135
feedLabeflag.pack(side=LEFT)
feedLabeflag.place(x=PlaceLable2_X,y=PlaceLable_Y)

##################### material #####################

# material Details label
materialLabel = Label(materialSettingsFrame, text = 'Material Details', font=sub_label_font,fg=label_color)
# material_Lable_place_x=80
# material_Lable_place_y=60

materialLabel.pack(side=LEFT)
materialLabel.place(x=Lable_place_x,y=Lable_place_y)

###material tabel###
scroll_width=20
tab_side=LEFT
x=250
y= 150
frame=materialSettingsFrame
list_height=5
# table_place_x = 80
# table_place_y=80

columns_name_list=['    Material   ']

queryMaterial = "SELECT * FROM material WHERE ISNULL(deleted)"

material_tabel=table(frame,scroll_width,list_height,tab_side,x,y,table_place_x,
                   table_place_y)
material_tabel.create_fully_tabel( columns_name_list, queryMaterial)

###material functions###
def editMaterialfun():
    selected_rec = material_tabel.selected()
    selected_non = material_tabel.selected_is_non(selected_rec)
    if not selected_non:
        # popup_size = "800x450"
        popup_size = "700x550"
        editMaterialPopup = Popup()
        editMaterialPopup.open_pop('Edit Material Details', popup_size)

        query = "UPDATE material SET materialName = %s   WHERE idmaterial = %s"
        table_name = 'material'
        pk = selected_rec[1]
        labels = [('Material', '')]
        save_title = "Save Changes"

        editMaterialPopup.edit_popup(labels, selected_rec, save_title, query, pk, material_tabel,  table_name)


def addMaterialfun():
    addMaterialPopup = Popup()
    # popup_size = "800x450"
    popup_size = "900x550"
    addMaterialPopup.open_pop('Add Material Details', popup_size)
    labels = [('Material', '')]
    save_title = "Add Material"
    insetQuery = "INSERT INTO material SET materialName = %s "
    table_name='material'
    addMaterialPopup.add_popup(labels, save_title, insetQuery, material_tabel, table_name)

def deleteMaterialfun():
    query = "DELETE FROM material WHERE idmaterial = %s"
    table_name='material'
    material_tabel.delete_record(query,table_name)


#material buttons

#Create a button in the main Window to add record - material
materialAddIcon = Image.open("addIcon.png")
resizedMaterialAddIcon = materialAddIcon.resize((25, 25), Image.ANTIALIAS)
imgAddMaterial = ImageTk.PhotoImage(resizedMaterialAddIcon)
addMaterialButton = Button(materialSettingsFrame, image=imgAddModule, borderwidth=0, command=addMaterialfun)
addMaterialButton.pack(side= LEFT)
addMaterialButton.place(x=table_place_x + material_tabel.winfo_reqwidth() - 70, y=table_place_y+20)

#Create a button in the main Window to edit  record (open the popup) - material
materialEditIcon = Image.open("editIcon.jpg")
resizedMaterialEditIcon = materialEditIcon.resize((20, 20), Image.ANTIALIAS)
imgEditMaterial = ImageTk.PhotoImage(resizedMaterialEditIcon)
editMaterialButton = Button(materialSettingsFrame, image=imgEditMaterial, borderwidth=0, command=editMaterialfun)
editMaterialButton.pack(side= LEFT)
editMaterialButton.place(x=table_place_x + material_tabel.winfo_reqwidth() - 30, y=table_place_y+22)

#Create a button in the main Window to Delete record - material
materialDeleteIcon = Image.open("‏‏deleteIcon.png")
resizedMaterialDeleteIcon = materialDeleteIcon.resize((20, 20), Image.ANTIALIAS)
imgDeleteMaterial = ImageTk.PhotoImage(resizedMaterialDeleteIcon)
deleteMaterialButton = Button(materialSettingsFrame, image=imgDeleteMaterial, borderwidth=0, command=deleteMaterialfun)
deleteMaterialButton.pack(side= LEFT)
deleteMaterialButton.place(x=table_place_x + material_tabel.winfo_reqwidth() +7, y=table_place_y+22)


##################### settings - Hospitals #####################
#hospital frame
hospitalFrame = Frame(root)
# hospitalFrame.pack(fill=X)

# feed label - hospital
feedLabel = Label(hospitalFrame, text = 'Settings ➝ ', font=label_font_flag,fg=label_color)
PlaceLable_X=50
PlaceLable_Y=10
feedLabel.pack(side=LEFT)
feedLabel.place(x=PlaceLable_X,y=PlaceLable_Y)

feedLabeflag = Label(hospitalFrame, text = 'hospital', font=label_font_flag_on_page,fg=label_color)

PlaceLable2_X=135
feedLabeflag.pack(side=LEFT)
feedLabeflag.place(x=PlaceLable2_X,y=PlaceLable_Y)


# hospital Details label
hospitalLabel = Label(hospitalFrame, text = 'Hospitals Details', font=sub_label_font,fg=label_color)
# module_Lable_place_x=80
# module_Lable_place_y=60

hospitalLabel.pack(side=LEFT)
hospitalLabel.place(x=Lable_place_x,y=Lable_place_y)


#hospital table
scroll_width=20
tab_side=LEFT
x=895
y= 130
frame=hospitalFrame
list_height=30
c = 80

lable_place_x = 80
lable_place_y=70

columns_name_list=('        Name        ', 'Fixed Activity Level (mci)', 'Transport Time - min (minutes)', 'Transport Time - man (minutes)')

hospital_query="SELECT * FROM hospital WHERE ISNULL(deleted)"

hospital_tabel=table(frame,scroll_width,list_height,tab_side,x,y,lable_place_x,
                     lable_place_y)
hospital_tabel.create_fully_tabel( columns_name_list, hospital_query)

hospitalFrame.pack(fill='both',expand=1)

###hospital functions###
def editHospitalfun():
    selected_rec = hospital_tabel.selected()
    selected_non = hospital_tabel.selected_is_non(selected_rec)
    if not selected_non:
        editHospitalPopup = Popup()
        # popup_size = "800x450"
        popup_size = "900x550"
        editHospitalPopup.open_pop('Edit Hospital Details',popup_size)
        table_name= 'hospital'
        query = "UPDATE hospital SET Name = %s ,Fixed_activity_level= %s, Transport_time_min=%s ,Transport_time_max=%s WHERE idhospital = %s"

        pk = selected_rec[4]

        labels = (('Name', ''), ('Fixed activity level', '(mci/h)'),  ('Transport time - min', '(min)'),  ('Transport time - max', '(min)'))
        save_title = "Save Changes"

        editHospitalPopup.edit_popup(labels, selected_rec, save_title, query, pk, hospital_tabel,table_name)


def addHospitalfun():
    addHospitalPopup = Popup()
    # popup_size = "800x450"
    popup_size = "900x550"
    addHospitalPopup.open_pop('Add Hospital Details',popup_size)
    labels = (('Name', ''), ('Fixed activity level', '(mci/h)'), ('Transport time - min', '(min)'),('Transport time - max', '(min)'))
    save_title = "Add Hospital"
    insertQuery = "INSERT INTO hospital SET Name = %s ,Fixed_activity_level= %s,Transport_time_min=%s ,Transport_time_max=%s"
    # selectIDquery = "SELECT MAX(idhospital) FROM hospital"
    table_name = 'hospital'
    addHospitalPopup.add_popup(labels, save_title, insertQuery, hospital_tabel, table_name)

def deleteHospitalfun():
    query = "DELETE FROM hospital WHERE idhospital = %s"
    table_name= 'hospital'
    hospital_tabel.delete_record(query,table_name)

#hospital buttons

#Create a button in the main Window to add record - hospital
hospitalAddIcon = Image.open("addIcon.png")
resizedHospitalAddIcon = hospitalAddIcon.resize((25, 25), Image.ANTIALIAS)
imgAddHospital = ImageTk.PhotoImage(resizedHospitalAddIcon)
addHospitalButton = Button(hospitalFrame, image=imgAddHospital, borderwidth=0, command=lambda : addHospitalfun())
addHospitalButton.pack(side= LEFT)
addHospitalButton.place(x=lable_place_x + hospital_tabel.winfo_reqwidth() - 100, y=lable_place_y+14)


#Create a button in the main Window to edit  record (open the popup) - hospital
hospitalEditIcon = Image.open("editIcon.jpg")
resizedHospitalEditIcon = hospitalEditIcon.resize((20, 20), Image.ANTIALIAS)
imgEditHospital = ImageTk.PhotoImage(resizedHospitalEditIcon)
editHospitalButton = Button(hospitalFrame, image=imgEditHospital, borderwidth=0, command= lambda :editHospitalfun())

editHospitalButton.pack(side= LEFT)
editHospitalButton.place(x=lable_place_x + hospital_tabel.winfo_reqwidth() - 50, y=lable_place_y+15)


# Create a button in the main Window to Delete record - hospital
hospitalDeleteIcon = Image.open("‏‏deleteIcon.png")
resizedHospitalDeleteIcon = hospitalDeleteIcon.resize((20, 20), Image.ANTIALIAS)
imgDeleteHospital = ImageTk.PhotoImage(resizedHospitalDeleteIcon)
deleteHospitalButton = Button(hospitalFrame, image=imgDeleteHospital, borderwidth=0, command=lambda : deleteHospitalfun())
deleteHospitalButton.pack(side=LEFT)

deleteHospitalButton.place(x=lable_place_x + hospital_tabel.winfo_reqwidth() , y=lable_place_y + 15)

#################### Work Plan Page #####################
#Work Plan frame
WorkPlanFrame = Frame(root)
# h = Scrollbar(WorkPlanFrame, orient='horizontal')
WorkPlanFrame.pack(fill=X)

##################### Work Plan #####################
# Work Plan Details label
WorkPlanLabel = Label(WorkPlanFrame, text = 'Work Plans', font=sub_label_font,fg=label_color)
Lable_place_x=80
Lable_place_y=60

WorkPlanLabel.pack(side=LEFT)
WorkPlanLabel.place(x=Lable_place_x,y=Lable_place_y)

###Work Plan tabel###
scroll_width=20
tab_side=LEFT
x=310
y= 140
frame=WorkPlanFrame
list_height=50
table_place_x = 80
table_place_y = 80
columns_name_list=('    Date   ',' Material ' )
query = "SELECT WP.idworkplan, WP.Date, m.materialName FROM workplan WP JOIN material M ON WP.materialID=M.idmaterial WHERE ISNULL(WP.deleted) "

wp_tabel=table(frame,scroll_width,list_height,tab_side,x,y,table_place_x,
                  table_place_y,)
wp_tabel.create_fully_tabel( columns_name_list, query)


def show_wp(evet):
    selected_rec = wp_tabel.selected()
    pk = selected_rec[2]
    query_hospital = """SELECT DISTINCT(b.idbatch + b.batchNumber),b.batchNumber, h.Name
            FROM batch b 
            LEFT JOIN orders o ON o.batchID = b.idbatch
            JOIN hospital h ON h.idhospital = o.hospitalID
             WHERE b.workplanID = """ + str(pk) + " ORDER BY b.batchNumber"
    cursor.execute(query_hospital)
    data_h = cursor.fetchall()
    show_wp_popup = Popup()

    query_batches = """SELECT b.idbatch , b.batchNumber,b.EOS_TIME, b.Total_eos
                FROM batch b WHERE b.workplanID = """ + str(pk) + " ORDER BY b.batchNumber"
    cursor.execute(query_batches)
    data_b = cursor.fetchall()

    title = selected_rec[0] + '  |  ' + selected_rec[1]
    geo = "800x450"
    show_wp_popup.open_pop(title, geo)

    res_table = table(show_wp_popup, 20,15,LEFT,505,100,90,30)

    columns_name_list = ('    #    ','  Batch 1  ', '  Batch 2  ', '  Batch 3  ')
    res_table['columns'] = columns_name_list

    res_table.column("#0", width=0, stretch=NO)
    res_table.heading("#0", text="", anchor=CENTER)

    i = 0
    len_of_col = len(columns_name_list)
    for column_name in columns_name_list:
        # column format
        if i == 0 or i == len_of_col - 2:
            width = len(column_name) * 6 + 30
        else:
            width = len(column_name) * 6

        res_table.column(column_name, anchor=CENTER, width=width)
        # # Create Headings
        res_table.heading(column_name, text=column_name, anchor=CENTER)

    val_eos = ['EOS Time']
    val_activity = ['Activity']
    iid_eos=0
    iid_activity=1
    for recorf in data_b:
        val_eos.append(recorf[2])
        val_activity.append(recorf[3])

    res_table.insert(parent='', index='end', iid=iid_eos, text='',
                    values=val_eos)
    res_table.insert(parent='', index='end', iid=iid_activity, text='',
                     values=val_activity)
    res_table.pack()

    val = ['Hospitals']
    batch=[]
    h_b=""
    iid_hospitals = 2
    for recor in data_h:

        if recor[1] in batch:
            h_b += recor[2] +'\n'
            if recor==data_h[len(data_h)-1]:  #if its the last record (for append the hospitals to the list)
                val.append(h_b)
        else:
            if len(batch)!=0:
                val.append(h_b)
            h_b = recor[2] +'\n'
            batch.append(recor[1])

    res_table.insert(parent='', index='end', iid=iid_hospitals, text='',
                     values=val)
    res_table.pack()


wp_tabel.bind('<Double-1>', show_wp)


# ###Work Plan functions###
# def editWPfun():
#     selected_rec = cyclo_tabel.selected()
#     selected_non=cyclo_tabel.selected_is_non(selected_rec)
#     if not selected_non:
#         editCyclPopup = Popup()
#         editCyclPopup.open_pop('Edit Cyclotron Details')
#
#         query = "UPDATE resourcecyclotron SET version = %s ,capacity= %s, constant_efficiency= %s,description=%s  WHERE idresourceCyclotron = %s"
#         pk = selected_rec[4]
#         table_name = 'resourcecyclotron'
#         labels = (('Version', ''), ('Capacity', '(mci/h)'), ('Constant Efficiency', '(mCi/mA)'), ('Description', ''))
#         save_title = "Save Changes"
#
#         editCyclPopup.edit_popup(labels, selected_rec, save_title, query, pk, cyclo_tabel,table_name)
#
#
def deleteWPfun():
    query = "DELETE FROM workplan WHERE idworkplan = %s"
    table_name='workplan'
    wp_tabel.delete_WP_record()

def addWPfun():
    addWPPopup = Popup()
    popup_size = "800x450"
    addWPPopup.open_pop('Create Work Plan',popup_size)
    addWPPopup.add_wp_popup()

# #work plan buttons
# #Create a button in the main Window to edit  record (open the popup) - work plan
# wpEditIcon = Image.open("editIcon.jpg")
# resizedWPEditIcon = wpEditIcon.resize((20, 20), Image.ANTIALIAS)
# imgEditwp = ImageTk.PhotoImage(resizedWPEditIcon)
# # editCyclotronButton = Button(ctcloSettingsFrame, image=imgEditCyclotron, borderwidth=0, command= lambda :editCyclotronfun())
# editWPButton = Button(WorkPlanFrame, image=imgEditwp, borderwidth=0, command= lambda :editCyclotronfun())
#
# editWPButton.pack(side= LEFT)
# editWPButton.place(x=table_place_x+450, y=table_place_y+15)
#
#Create a button in the main Window to add record - work plan
wpAddIcon = Image.open("addIcon.png")
resizedWPAddIcon = wpAddIcon.resize((25, 25), Image.ANTIALIAS)

imgAddWP = ImageTk.PhotoImage(resizedWPAddIcon)
addWPButton = Button(WorkPlanFrame, image=imgAddWP, borderwidth=0, command=lambda : addWPfun())
addWPButton.pack(side= LEFT)
addWPButton.place(x=table_place_x + wp_tabel.winfo_reqwidth() - 45, y=table_place_y+15)


# Create a button in the main Window to Delete record - work plan
wpDeleteIcon = Image.open("‏‏deleteIcon.png")
resizedWPDeleteIcon = wpDeleteIcon.resize((20, 20), Image.ANTIALIAS)
imgDeleteWP = ImageTk.PhotoImage(resizedWPDeleteIcon)
deleteWPButton = Button(WorkPlanFrame, image=imgDeleteWP, borderwidth=0, command=lambda : deleteWPfun())
deleteWPButton.pack(side=LEFT)
deleteWPButton.place(x=table_place_x + wp_tabel.winfo_reqwidth() , y=table_place_y+15)

################### batches #################
#################### batch Page #####################
#batch frame
batchFrame = Frame(root)
# h = Scrollbar(batchFrame, orient='horizontal')
batchFrame.pack(fill=X)

# batch Details label
BatchLabel = Label(batchFrame, text = 'Batches', font=sub_label_font,fg=label_color)
Lable_place_x=80
Lable_place_y=60

BatchLabel.pack(side=LEFT)
BatchLabel.place(x=Lable_place_x,y=Lable_place_y)

#batches table
scroll_width=20
tab_side=LEFT
x=1050
y= 130
frame=batchFrame
list_height=30
c = 80

lable_place_x = 80
lable_place_y=70

columns_name_list=('  Date  ','Material', 'Batch Number','Time leaves Hadassah','Total EOS (mCi)',' EOS Time ','TargetCurrentLB ', 'DecayCorrected_TTA (mCi)')

batch_query="""SELECT  b.idbatch , wp.Date ,m.materialName,b.batchNumber,b.Time_leaves_Hadassah,b.Total_eos,b.EOS_TIME, b.TargetCurrentLB ,b.DecayCorrected_TTA
                FROM batch b 
                JOIN workplan wp ON wp.idworkplan = b.workplanID 
                JOIN material m ON m.idmaterial = wp.materialID"""

batch_tabel=table(frame,scroll_width,list_height,tab_side,x,y,lable_place_x,
                     lable_place_y)
batch_tabel.create_fully_tabel( columns_name_list, batch_query)

batchFrame.pack(fill='both',expand=1)


###batch functions###
def editBatchfun():
    selected_rec = batch_tabel.selected()
    selected_non = batch_tabel.selected_is_non(selected_rec)
    if not selected_non:
        editBatchPopup = Popup()
        # popup_size = "800x450"
        popup_size = "900x570"
        editBatchPopup.open_pop('Edit Batch Details',popup_size)
        table_name= 'batch'
        query = "UPDATE batch SET Time_leaves_Hadassah=%s,Total_eos=%s ,EOS_TIME = $s,TargetCurrentLB = %s ,DecayCorrected_TTA= %s  WHERE idbatch = %s"

        pk = selected_rec[7]

        labels = ( ('Time leaves Hadassah',''),('Total EOS', '(mCi/h)'),('EOS Time',''),('TargetCurrentLB', ''), ('Decay Corrected TTA', '(mCi/h)'))
        save_title = "Save Changes"

        # def edit_popup(self, labels, valueList, save_title, *args, table_name):
        editBatchPopup.edit_popup(labels, selected_rec, save_title, query, pk, batch_tabel,table_name)

#batch buttons

#Create a button in the main Window to edit  record (open the popup) - hospital
batchEditIcon = Image.open("editIcon.jpg")
resizedBatchEditIcon = batchEditIcon.resize((20, 20), Image.ANTIALIAS)
imgEditBatch = ImageTk.PhotoImage(resizedBatchEditIcon)
editBatchButton = Button(batchFrame, image=imgEditBatch, borderwidth=0, command= lambda :editBatchfun())

editBatchButton.pack(side= LEFT)
editBatchButton.place(x=lable_place_x + batch_tabel.winfo_reqwidth() - 50, y=lable_place_y+15)


cycloSettingsFrame.forget()
moduleSettingsFrame.forget()
hospitalFrame.forget()
WorkPlanFrame.forget()
batchFrame.forget()

root.mainloop()