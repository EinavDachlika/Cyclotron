from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import datetime
from PIL import Image, ImageTk, ImageFont
import mysql.connector
from mysql.connector import Error
from datetime import datetime , date, timedelta
# import xlrd
import xlwt
import openpyxl
from openpyxl import *
from openpyxl.styles import *
from openpyxl.utils import get_column_letter
from operator import itemgetter


import math
from numpy import log as ln
import numpy as np

import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.patches import Patch


#
# # Declaring a figure "gnt"
# fig, gnt = plt.subplots()
#
# # Setting Y-axis limits
# gnt.set_ylim(0, 50)
#
# # Setting X-axis limits
# gnt.set_xlim("07:00", "21:00")
#
# # Setting labels for x-axis and y-axis
# gnt.set_xlabel('Time')
# gnt.set_ylabel('Batch')
#
# # Setting ticks on y-axis
# gnt.set_yticks([15, 25, 100])
# # Labelling tickes of y-axis
# gnt.set_yticklabels(['Batch 1', 'Batch 2', 'Batch 3'])
#
# # Setting graph attribute
# gnt.grid(True)
#
# # Declaring a bar in schedule
# gnt.broken_barh([(40, 50)], (30, 9), facecolors=('tab:orange'))
#
# # Declaring multiple bars in at same level and same width
# gnt.broken_barh([(110, 10), (150, 10)], (10, 9),
#                 facecolors='tab:blue')
#
# gnt.broken_barh([(10, 50), (100, 20), (130, 10)], (20, 9),
#                 facecolors=('tab:red'))

# fig, ax = plt.subplots(1, figsize=(16,6))
# ax.barh(('Batch 3','Batch 2', 'Batch 1'), (datetime.strptime('07:00:00', '%H:%M:%S').strftime('%H:%M'),datetime.strptime('21:00:00', '%H:%M:%S').strftime('%H:%M'),datetime.strptime('21:00:00', '%H:%M:%S').strftime('%H:%M'),), left=(0,4,5), color=('blue','orange','black'))
# ##### LEGENDS #####
# c_dict = {'MKT':'#E64646', 'FIN':'#E69646', 'ENG':'#34D05C',
#           'PROD':'#34D0C3', 'IT':'#3475D0'}
# legend_elements = [Patch(facecolor=c_dict[i], label=i)  for i in c_dict]
# plt.legend(handles=legend_elements)
#
# ##### TICKS #####
# xticks = np.arange(0, 24, 3)
# xticks_labels = pd.date_range(datetime.strptime('15:00:00', '%H:%M:%S').strftime('%H:%M'), end=datetime.strptime('15:00:00', '%H:%M:%S').strftime('%H:%M')).strftime("%h:%m")
# # xticks_minor = np.arange(0, 15, 1)
# ax.set_xticks(xticks)
# # ax.set_xticks(xticks_minor, minor=True)
# # ax.set_xticklabels(xticks_labels[::3])
#
# plt.show()

import plotly_express as px
import plotly.graph_objects as go
#
# import pandas as pd
# # fig, ax = plt.subplots(1, figsize=(5,18))
# fig, gnt = plt.subplots()
#
#
# df = pd.DataFrame([
#     # dict(Batch="Batch 1",amount=3, Start=datetime.strptime('05:00:00', '%H:%M:%S').strftime('%H:%M'), Finish=datetime.strptime('14:00:00', '%H:%M:%S').strftime('%H:%M'), Hospital="Hadassa"),
#     # dict(Batch="Batch 1",amount=3, Start='10:30', Finish='14:30', Hospital="Max"),
#     dict(Batch="Batch 2",amount=3, Start='05:30', Finish='20:30', Hospital="Alex"),
#     dict(Batch="Batch 3",amount=3, Start='21:30', Finish='02:30', Hospital="Max")
# ])
# fig = px.timeline(df, x_start="Start", x_end="Finish", y="Batch", color="Hospital")
# fig.update_yaxes(autorange="reversed")
#
# # fig.show()
#
# #############3
# fig, ax = plt.subplots(1, figsize=(16,6))
# ax.barh(df.Batch,df.amount , left=df.Start, color=df.Hospital)


##############3



import plotly.graph_objects as go # or plotly.express as px
# fig = go.Figure() # or any Plotly Express function e.g. px.bar(...)
# fig.add_trace( ... )
# fig.update_layout( ... )
#
# import dash
# from dash import dcc, Dash
# from dash import html
#
# app = dash.Dash()
# app.layout = html.Div([
#     dcc.Graph(figure=fig)
# ])
#
# app.run_server(debug=True, use_reloader=False)  # Turn off reloader if inside Jupyter
# # from xlwt import Workbook
# # from tkcalendar import Calendar,DateEntry
# # import csv
# # from importlib import reload
#
# app = Dash('Einav-test')
#
# colors = {
#     'background': '#111111',
#     'text': '#7FDBFF'
# }
#
# # assume you have a "long-form" data frame
# # see https://plotly.com/python/px-arguments/ for more options
# df = pd.DataFrame({
#     "Fruit": ["Apples", "Oranges", "Bananas", "Apples", "Oranges", "Bananas"],
#     "Amount": [4, 1, 2, 2, 4, 5],
#     "City": ["SF", "SF", "SF", "Montreal", "Montreal", "Montreal"]
# })
#
# fig = px.bar(df, x="Fruit", y="Amount", color="City", barmode="group")
#
# fig.update_layout(
#     plot_bgcolor=colors['background'],
#     paper_bgcolor=colors['background'],
#     font_color=colors['text']
# )
#
# app.layout = html.Div(style={'backgroundColor': colors['background']}, children=[
#     html.H1(
#         children='Hello Dash',
#         style={
#             'textAlign': 'center',
#             'color': colors['text']
#         }
#     ),
#
#     html.Div(children='Dash: A web application framework for your data.', style={
#         'textAlign': 'center',
#         'color': colors['text']
#     }),
#
#     dcc.Graph(
#         id='example-graph-2',
#         figure=fig
#     )
# ])
#
# if __name__ == '__main__':
#     app.run_server(debug=True)

root = Tk()
#root.geometry("300x300")

root.title("Settings")

#defult font
root.option_add("*Font", "Helvetica")


# connect to MySqL
try:

    # # Maor local DB Mysql
    # db = mysql.connector.connect(
    #     host="localhost",
    #     port=3308,
    #     user="root",
    #     password="root",
    #     database="cyclotron")

    # Einav local DB-Mysql
    db = mysql.connector.connect(
      host="localhost",
      user="root",
      password="Cyclotron2022@?%",
      database= "cyclotron")

    if db.is_connected():
        # db_Info = db.get_server_info()
        # print("Connected to MySQL Server version ", db_Info)
        dbCursor = db.cursor(buffered=True)
        # dbCursor.execute("select database();")
        # record = dbCursor.fetchone()
        # print("You're connected to database: ", record)

except Error as e:
    print("Error while connecting to MySQL", e)

# excelIcon = Image.open("excelIcon.png")
# resizedExcelIcon = excelIcon.resize((20, 20), Image.ANTIALIAS)
# imgExcel = ImageTk.PhotoImage(resizedExcelIcon)
# ExcelButton = Button(root, image=imgExcel, borderwidth=0,
#                              command=lambda: root.export_WP_To_Excel())
# # ExcelButton.pack(side=LEFT)
# ExcelButton.place(x= 80, y=80)



cursor = db.cursor()
date= date(2022,8,2)
# print(date)
# query = """SELECT h.Name,h.Fixed_activity_level , o.injection_time,o.amount, m.materialName, o.Date
#         FROM hospital h INNER JOIN orders o ON  h.idhospital=o.hospitalID INNER JOIN material m ON m.idmaterial=o.materialID
#         where Date = '""" + str(date)+""" ' and o.materialID=1
#         ORDER BY hospitalID, injection_time """
# # print(query)
# # query = "SELECT Date FROM orders "
#
# cursor.execute(query)
# data = cursor.fetchall()
# print(data)
# for order in data:
#     print(order)

#
# def exportCsv(result):
#     """Function for creating/exporting Excel file"""
#     print("try exporting new excel file...");
#     headers = ['OrderId', 'Date', 'Injection Time', 'Amount','idhospital','batchID','decayCorrected'];
#     with open('orders.csv','a',newline="") as f:
#         w = csv.writer(f,dialect='excel');
#         messagebox.showinfo("message","Excel file was created");
#         # write the headers
#         w.writerow(headers);
#         for record in result:
#             w.writerow(record);
#
# exportCsv(['OrderId', 'Date', 'Injection Time', 'Amount','idhospital','batchID','decayCorrected'])

# def exportExcel(data):
#     wb = openpyxl.Workbook()
#     # wb.add_format({'text_wrap': True})
#     sheet = wb.active
#     wb.font = Font(size=36, bold=True)
#
#     yellow = "ffff99"
#     #headers
#     headers = ["Dose","Batch","Dose mCi","EOS time", "Cal. Time", "Injection Time", "Decay Time GMP","Decay corrected GMP" ]
#     headers_col_i = 4
#     for header in headers:
#         sheet.cell(row=6, column=headers_col_i).value = header
#         # sheet.cell(row=6, column=headers_col_i).font = Font(size=20, bold=True)
#         sheet.cell(row=6, column=headers_col_i).fill = PatternFill(start_color=yellow, end_color=yellow,
#                                     fill_type="solid")  # bg of hospital cell
#         headers_col_i+=1
#
#     headers2 = ["TIME LEAVES HADASSAH","PRODCUTION SITE","NUMBER OF HOSPITAL VIALS (Not including Hadassah", "TIME USED FOR CALIBRATION", "EOS" ]
#     headers_row_i = 1
#
#     for header in headers2:
#         sheet.cell(row=headers_row_i, column=4).value =header
#         sheet.merge_cells(start_row=headers_row_i, start_column=4, end_row=headers_row_i,
#                                                       end_column=10)
#         headers_row_i+=1
#
#
#     hospitals=[]
#     row_index = 9
#     print(data)
#     for order in data:
#         if order[0] not in hospitals:
#
#             i = 1
#             col_index = 3
#             hospital_orders = [row[1:] for row in data if row[0] == order[0]]
#             end_row_to_merge = row_index+ len(hospital_orders) -1
#
#             hospital_name_cell = sheet.cell(row=row_index, column=col_index)
#             hospital_name_cell.value = order[0] # insert hoapital name to the first col
#             merge_hospital_name_cells = sheet.merge_cells(start_row=row_index, start_column=col_index, end_row=end_row_to_merge, end_column=col_index)
#             wrap_alignment = Alignment(wrap_text=True)
#             hospital_name_cell.alignment = wrap_alignment
#             # hospital_name_cell.font = Font(size=35, bold=True)
#             grey = "c0c0c0"
#             hospital_name_cell.fill = PatternFill(start_color=grey, end_color=grey,
#                                     fill_type="solid")  # bg of hospital cell
#             hospitals.append(order[0])
#             print(hospital_orders)
#             for row in hospital_orders:
#                 DosemCi = row[0] * row[2]
#                 sheet.cell(row=row_index, column=4).value = i  #serial number
#                 # sheet.cell(row=row_index, column=4).font= Font(size=36, bold=True)
#
#                 sheet.cell(row=row_index, column=6).value = DosemCi
#
#                 sheet.cell(row=row_index, column=9).value = str(row[1])   #injection time
#                 # sheet.cell(row=row_index, column=9).font = Font(size=36, bold=True)
#                 i+=1
#                 row_index+=1
#             row_index += 1
#
#
#     wb.save('workplanExce1.xls')
# #
# # exportExcel(data)


# # FilePath = "FDG work plan template.xlsx"
# FilePath = "FDG format.xlsx"
#
# wb = load_workbook(FilePath)


# writer = pd.ExcelWriter(FilePath, engine = 'openpyxl')
# # writer.book = wb
# sheet = wb.active
# sheet=wb['work plan']
#
# hospitals = []
# row_index = 9
# for order in data:
#     if order[0] not in hospitals:
#         grey = "c0c0c0"
#         col_start = 4
#         col_end = 16
#
#         sheet.cell(row=row_index, column=col_start).fill = PatternFill(start_color=grey, end_color=grey,
#                                              fill_type="solid")  # bg of buffer cell
#
#         merge_buffer = sheet.merge_cells(start_row=row_index, start_column=col_start, end_row=row_index,
#                                          end_column=col_end)
#
#         i = 1
#         col_index = 3
#         row_index += 1
#         hospital_orders = [row[1:] for row in data if row[0] == order[0]]
#         end_row_to_merge = row_index + len(hospital_orders) - 1
#         hospital_name_cell = sheet.cell(row=row_index, column=col_index)
#         hospital_name_cell.value = order[0]  # insert hoapital name to the first col
#         merge_hospital_name_cells = sheet.merge_cells(start_row=row_index, start_column=col_index,
#                                                       end_row=end_row_to_merge, end_column=col_index)
#         hospitals.append(order[0])
#
#         for row in hospital_orders:
#             DosemCi = row[0] * row[2]
#             # sheet.cell(row=row_index, column=4).value = i  # serial number
#             # sheet.cell(row=row_index, column=6).value = DosemCi
#             # sheet.cell(row=row_index, column=11).value = str(row[1])  # injection time
#
#             sheet.cell(row=row_index, column=4).value = i  # serial number
#             sheet.cell(row=row_index, column=6).value = DosemCi
#             sheet.cell(row=row_index, column=9).value = str(row[1])  # injection time
#             i += 1
#             row_index += 1


# wb.save('workplanExce020822.xls')


# df = pd.DataFrame([[1,2,3,4,5,6,7], [8,9,10,11,12,13,14],[15,16,17,18,19,20,21]], index=[11,12,13,14,15,16,17],columns=['d', 'e', 'f',',g','h','i','g'])
# df.to_excel(FilePath, sheet_name = 'work plan')
# writer.save()
# writer.close()



#1.9.22
# algorithm functions
lamda = ln(2) / 109.6
max_activity_batch = 7300
def sortByTout(hospital):
    return hospital["tout_required"]

def sortByToutActually(hospital):
    return hospital["Tout_actually"]

def sortByTeos(hospital):
    return hospital["eos_req"]

def flat_list(batches):
    flat_list = []
    batch_num=0
    for b in batches:
        batch_num+=1
        if not len(b)==0:
            for order in b:
                order['batch'] = batch_num
                flat_list.append(order)
    return flat_list

def recursion_for_tout(hospital_data):
    min_to_add = 15
    add_to_teos = timedelta(minutes=min_to_add)
    first_teos = hospital_data[0]["eos_req"]
    for h in hospital_data:
        Tout_actually = first_teos + add_to_teos
        if Tout_actually > h["tout_required"]:
                diff = Tout_actually - h["tout_required"]
                diff_to_int = diff.seconds / 60
                Subtract_from_eos = timedelta(minutes=diff_to_int)
                last_eos = hospital_data[0]["eos_req"]
                updated_eos = last_eos - Subtract_from_eos
                hospital_data[0]["eos_req"] = updated_eos
                recursion_for_tout(hospital_data)

        h["Tout_actually"] = Tout_actually
        min_to_add += 5
        add_to_teos = timedelta(minutes=min_to_add)

def main_algorithm_calculation(batches, hospitals_output, batches_general_data, data_original):  # test - Einav

    data=data_original
    hospital_data = []
    index = 1
    for i in range(0,3):
        hospitals = []
        batch = []
        for order in data:
            order_time = datetime.strptime(str(order["injection_time"]), '%H:%M:%S').time()
            hospital_name = order['Name']
            if hospital_name not in hospitals:  # order[1] = hospital name in order
                hospitals.append(hospital_name)
                # add hospital record to hospitals_output list
                try:
                    hospital = next(h for h in hospitals_output if h["Name"] == order["Name"] and h["Batch"] == index)
                except:
                    hospitals_output.append({"Name": order["Name"], "Activity": 0, 'Batch': index})

            if index != 1:  # condition for choosing Transport_time (min/max) - according to number of batch
                tout_temp = order['injection_time'] - timedelta(minutes=(order['Transport_time_max']))  # T1-Transport_time
                hospital_data.append(
                    {'hospital_name': hospital_name, 'batch': index, "injection_time": order['injection_time'],
                     "Transport_time": order['Transport_time_max'], "tout_required": tout_temp})

            else:
                tout_temp = order['injection_time'] - timedelta(minutes=(order['Transport_time_min']))  # T1-Transport_time
                hospital_data.append(
                    {'hospital_name': hospital_name, 'batch': index, "injection_time": order['injection_time'],
                     "Transport_time": order['Transport_time_min'], "tout_required": tout_temp})
            hospital_data.sort(key=sortByTout)  # sort hospital_data by tout_temp

            # Hospital_delivery_order = 1
            # interval_5 = 15
            # minutes_for_eos_cal = timedelta(minutes=interval_5)
            # for h in hospital_data:
            #     # delivery_order
            #     h["delivery_order"] = Hospital_delivery_order
            #     hospital_record = next(hospital for hospital in hospitals_output if
            #                            hospital["Name"] == h['hospital_name'] and hospital['Batch'] == h['batch'])
            #     hospital_record['delivery_order'] = Hospital_delivery_order
            #     Hospital_delivery_order += 1
            #
            #     eos_req = h["tout_required"] - minutes_for_eos_cal  # tout - intervals consider the order
            #     h["eos_req"] = (eos_req)
            #     interval_5 += 5
            #     minutes_for_eos_cal = timedelta(minutes=interval_5)
            #     # hospital_data.sort(key=sortBy)  # sort hospital_data by tout_eos
            #
            # hospital_data.sort(key=sortByTeos)  # sort hospital_data by tout_eos
            # # print(hospital_data)
            # recursion_for_tout(hospital_data)
            #
            # # save tout actual for each hospital
            # for h in hospital_data:
            #     hospital_record = next(hospital for hospital in hospitals_output if
            #                            hospital["Name"] == h['hospital_name'] and hospital['Batch'] == h['batch'])
            #     hospital_record['Tout_actually'] = h['Tout_actually']
            #     # print(h["hospital_name"] , " actually: ", h['Tout_actually'])
            #
            # # insert Teos - new module
            # Teos_key = "Teos"
            # t_eos_final = hospital_data[0][
            #     "eos_req"]  # first index in the list is the shortest time of eos (because it's sorted)
            # batches_general_data[index][Teos_key] = t_eos_final
            #
            # # insert Tout
            # Tout_key = "Tout"
            # hospital_data.sort(key=sortByToutActually)
            # last_index = len(hospital_data) - 1
            # t_out_final = hospital_data[last_index]['Tout_actually']  # last tout actually
            # batches_general_data[index][Tout_key] = t_out_final
            #
            # # insert Tcal - new module
            # Tcal_key = "Tcal"
            # minutes_for_Tcal_cal = timedelta(minutes=30)
            # t_cal = t_out_final + minutes_for_Tcal_cal
            # batches_general_data[index][Tcal_key] = t_cal
            #
            # bottles_b = len(hospitals)
            # bottels_key = "bottles_mum"
            # batches_general_data[index][bottels_key] = bottles_b
            #
            # batches_general_data[index]["Activity"] = 0  # define the key
            # for order in data:
            #     # insert (A) Activity_Tcal - Activity for Tcal time
            #     A_Tcal_key = "Activity_Tcal"
            #     injection_time = order["injection_time"]
            #     T_cal = batches_general_data[index]["Tcal"]
            #     A = order["Fixed_activity_level"]
            #
            #     diff = (injection_time - T_cal).total_seconds() / 60  # convert to minutes and then to float
            #     A_Tcal = math.ceil(A * math.pow((math.e), diff * lamda))  # math.ceil is round up
            #
            #     if batches_general_data[index]["Activity"] + A_Tcal >= max_activity_batch:
            #         batches[index + 1].append(order)
            #         batches[index].remove(order)

            ##         main_algorithm_calculation(batches, hospitals_output, batches_general_data)
            #
            #     else:
            #
            #         hospital = next(h for h in hospitals_output if h["Name"] == order["Name"] and h["Batch"] == index + 1)
            #         order[A_Tcal_key] = A_Tcal
            #         order['diff_Tcal_injectionT'] = diff
            #         hospital["Activity"] += A_Tcal
            #
            #         batch.append(order)
            #
            #         batches_general_data[index]["Activity"] += A_Tcal * 1.05
        batches.append(batch)
        index += 1



#
selected_date="2022-08-02"
selected_material="FDG"
 # algorithm
query = """SELECT o.idorders, h.Name,o.DoseNumber,h.Fixed_activity_level*o.amount as Fixed_activity_level, o.injection_time,o.amount,h.Transport_time_min,h.Transport_time_max
        FROM hospital h INNER JOIN orders o ON  h.idhospital=o.hospitalID INNER JOIN material m ON m.idmaterial=o.materialID
        where Date = '""" + str(selected_date) + """ ' and m.materialName= '""" +str(selected_material)+ """' ORDER BY injection_time """

cursor = db.cursor(dictionary=True)
cursor.execute(query)
data = cursor.fetchall()
print('data: ', data)
cursor = db.cursor(dictionary=False)

batches = []
dict_batch1_general = {}
dict_batch2_general = {}
dict_batch3_general = {}
batches_general_data = [dict_batch1_general, dict_batch2_general, dict_batch3_general]
hospitals_output = []  # for output
# main_algorithm_calculation(batches, hospitals_output, batches_general_data)
main_algorithm_calculation(batches, hospitals_output, batches_general_data, data)
# hospitals_output.sort(key=lambda hb:(hb['Batch'],hb['delivery_order']))
print("batches: ",batches)
print("hospitals_output: ",hospitals_output)
print("batches_general_data: ",batches_general_data)
all_batches_output = flat_list(batches)
all_batches_output.sort(key=itemgetter('Name'))
cursor = db.cursor()



# root.mainloop()